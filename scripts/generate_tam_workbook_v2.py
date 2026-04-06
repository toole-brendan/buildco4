#!/usr/bin/env python3
"""
Generate the Top-Down TAM Workbook for In-Service Vessel Work (US Navy & Coast Guard).

Styling follows docs/WORKBOOK_STYLING_GUIDE.md:
  - Arial 8 everywhere, row height 10.0, no gridlines, no freeze panes, no text wrap
  - Blue text = hardcoded numeric input
  - Black text = formula or text label
  - Green text = cross-tab link
  - Gray text = notes, sources, memos
  - Percentage cells always italic
"""

import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.worksheet.views import SheetView, SheetViewList
from openpyxl.utils import get_column_letter
from copy import copy


# ── Page-map builder: line number → PDF page ─────────────────────────────────

SOURCES_DIR = "topdown/sources"

def build_page_map(txt_filename):
    """Build a mapping from text-file line number to PDF page number.
    Uses form feed characters (\\f) inserted by pdftotext -layout."""
    path = os.path.join(SOURCES_DIR, txt_filename)
    page_map = {}
    pdf_page = 1
    try:
        with open(path, 'r', errors='replace') as fh:
            for i, line in enumerate(fh, 1):
                if '\f' in line:
                    pdf_page += 1
                page_map[i] = pdf_page
    except FileNotFoundError:
        pass
    return page_map

def fmt_page_ref(page_map, line_num, line_end=None):
    """Format a page reference string like 'p. 150' or 'pp. 150-155'."""
    if not page_map or not line_num:
        return ""
    p1 = page_map.get(line_num, None)
    if not p1:
        return ""
    if line_end:
        p2 = page_map.get(line_end, None)
        if p2 and p2 != p1:
            return f"pp. {p1}-{p2}"
    return f"p. {p1}"

# Build page maps for all source files
PAGE_MAPS = {}
for _tf in ['OMN_Book.txt', 'OPN_BA1_Book.txt', 'OPN_BA2_Book.txt',
            'OPN_BA4_Book.txt', 'OPN_BA5-8_Book.txt', 'SCN_Book.txt',
            'USCG_Justification.txt']:
    PAGE_MAPS[_tf] = build_page_map(_tf)

# Helper to get page ref from source file and line number
def pr(src_file, line_num, line_end=None):
    return fmt_page_ref(PAGE_MAPS.get(src_file, {}), line_num, line_end)

# ── Palette (8-char ARGB with FF alpha) ──────────────────────────────────────

BLACK       = "FF000000"
WHITE       = "FFFFFFFF"
NAVY        = "FF1F3150"
SLATE       = "FF5B7A99"
LIGHT_GRAY  = "FFF2F4F7"
INPUT_YELLOW = "FFFFF9C4"
CTRL_PURPLE = "FFEDE7F6"
KPI_TEAL    = "FFE0F2F1"
CAUTION_ORANGE = "FFFFF3E0"
ERROR_RED   = "FFFFEBEE"
PASS_GREEN  = "FFE8F5E9"
BLUE_INPUT  = "FF1F3150"
GRAY_TEXT   = "FF8C8C8C"
GREEN_LINK  = "FF2E7D32"
RED_EXT     = "FFC62828"

# ── Reusable styles ──────────────────────────────────────────────────────────

FONT_BASE        = Font(name="Arial", size=8, color=BLACK)
FONT_BOLD        = Font(name="Arial", size=8, color=BLACK, bold=True)
FONT_TITLE       = Font(name="Arial", size=8, color=WHITE, bold=True)
FONT_PURPOSE     = Font(name="Arial", size=8, color=GRAY_TEXT, italic=True)
FONT_BLUE        = Font(name="Arial", size=8, color=BLUE_INPUT)
FONT_BLUE_BOLD   = Font(name="Arial", size=8, color=BLUE_INPUT, bold=True)
FONT_BLUE_ITALIC = Font(name="Arial", size=8, color=BLUE_INPUT, italic=True)
FONT_GREEN       = Font(name="Arial", size=8, color=GREEN_LINK)
FONT_GREEN_BOLD  = Font(name="Arial", size=8, color=GREEN_LINK, bold=True)
FONT_GRAY        = Font(name="Arial", size=8, color=GRAY_TEXT)
FONT_GRAY_ITALIC = Font(name="Arial", size=8, color=GRAY_TEXT, italic=True)
FONT_BLACK_ITALIC = Font(name="Arial", size=8, color=BLACK, italic=True)
FONT_SECTION     = Font(name="Arial", size=8, color=WHITE, bold=True)
FONT_SUBSECTION  = Font(name="Arial", size=8, color=BLACK, bold=True)

FILL_WHITE      = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
FILL_BLACK      = PatternFill(start_color=BLACK, end_color=BLACK, fill_type="solid")
FILL_NAVY       = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
FILL_GRAY       = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
FILL_YELLOW     = PatternFill(start_color=INPUT_YELLOW, end_color=INPUT_YELLOW, fill_type="solid")
FILL_TEAL       = PatternFill(start_color=KPI_TEAL, end_color=KPI_TEAL, fill_type="solid")
FILL_ORANGE     = PatternFill(start_color=CAUTION_ORANGE, end_color=CAUTION_ORANGE, fill_type="solid")
FILL_RED        = PatternFill(start_color=ERROR_RED, end_color=ERROR_RED, fill_type="solid")
FILL_GREEN      = PatternFill(start_color=PASS_GREEN, end_color=PASS_GREEN, fill_type="solid")

ALIGN_LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=False)
ALIGN_RIGHT  = Alignment(horizontal="right", vertical="center", wrap_text=False)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)

BORDER_THIN_BOTTOM = Border(bottom=Side(style="thin", color=BLACK))
BORDER_THIN_TOP    = Border(top=Side(style="thin", color=BLACK))
BORDER_MED_TOP     = Border(top=Side(style="medium", color=BLACK))
BORDER_THICK_TOP   = Border(top=Side(style="thick", color=BLACK))

ROW_HEIGHT = 10.0
TAB_COLOR  = "1F3150"  # muted navy for all tabs

# Number format: zeros as dash, negatives in red parens
NUM_FMT_K   = '#,##0;[Red](#,##0);"-"'
NUM_FMT_PCT = '0.0%;[Red](0.0%);"-"'


# ── Helper functions ─────────────────────────────────────────────────────────

def setup_sheet(ws, title, purpose, col_count, tab_color=TAB_COLOR):
    """Apply standard header system and sheet settings."""
    ws.sheet_properties.tabColor = tab_color

    # Clean SheetView: no gridlines, no freeze panes
    sv_list = SheetViewList()
    sv_list.sheetView = [SheetView(showGridLines=False, workbookViewId=0)]
    ws.views = sv_list

    # Row 1: title band (black fill, white bold)
    for c in range(1, col_count + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = FILL_BLACK
        cell.font = FONT_TITLE
        cell.alignment = ALIGN_LEFT
    ws.cell(row=1, column=1).value = title

    # Row 2: purpose line (gray italic, white fill)
    for c in range(1, col_count + 1):
        cell = ws.cell(row=2, column=c)
        cell.fill = FILL_WHITE
        cell.font = FONT_PURPOSE
        cell.alignment = ALIGN_LEFT
    ws.cell(row=2, column=1).value = purpose

    # Row 3: spacer (white fill)
    for c in range(1, col_count + 1):
        cell = ws.cell(row=3, column=c)
        cell.fill = FILL_WHITE


def set_col_headers(ws, row, headers, col_count):
    """Write column header row with gray fill and thin bottom border."""
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_GRAY
        cell.font = FONT_BOLD
        cell.alignment = ALIGN_LEFT if c <= 2 else ALIGN_RIGHT
        cell.border = BORDER_THIN_BOTTOM
        if c <= len(headers):
            cell.value = headers[c - 1]


def white_row(ws, row, col_count):
    """Set explicit white fill on every cell in a row."""
    for c in range(1, col_count + 1):
        ws.cell(row=row, column=c).fill = FILL_WHITE


def write_data_row(ws, row, values, col_count, fonts=None, fills=None, indent=0, border=None):
    """
    Write a row of data.
    values: list of values by column
    fonts: optional dict of {col_index: Font}
    fills: optional dict of {col_index: PatternFill}
    """
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_WHITE
        cell.font = FONT_BASE
        cell.alignment = ALIGN_LEFT if c <= 2 else ALIGN_RIGHT
        if c <= len(values) and values[c - 1] is not None:
            cell.value = values[c - 1]
        if fonts and c in fonts:
            cell.font = fonts[c]
        if fills and c in fills:
            cell.fill = fills[c]
        if border:
            cell.border = border
    # Apply indent to col A
    if indent > 0:
        cell_a = ws.cell(row=row, column=1)
        cell_a.alignment = Alignment(horizontal="left", vertical="center",
                                     wrap_text=False, indent=indent)


def section_header(ws, row, text, col_count):
    """Black fill, white bold section header spanning all columns."""
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_BLACK
        cell.font = FONT_SECTION
        cell.alignment = ALIGN_LEFT
    ws.cell(row=row, column=1).value = text


def subsection_header(ws, row, text, col_count):
    """Gray fill, dark bold subsection header."""
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_GRAY
        cell.font = FONT_SUBSECTION
        cell.alignment = ALIGN_LEFT
    ws.cell(row=row, column=1).value = text


def set_row_heights(ws, max_row):
    """Set all rows to universal height."""
    for r in range(1, max_row + 1):
        ws.row_dimensions[r].height = ROW_HEIGHT


def set_col_widths(ws, widths):
    """widths: dict of {col_letter: width}"""
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


def fmt_k(cell, font=None):
    """Apply $K number format to a cell."""
    cell.number_format = NUM_FMT_K
    if font:
        cell.font = font


def fmt_pct(cell, font=None):
    """Apply percentage format (always italic)."""
    cell.number_format = NUM_FMT_PCT
    if not font:
        font = FONT_BLACK_ITALIC
    cell.font = font


# ── DATA ─────────────────────────────────────────────────────────────────────

# Bucket definitions
BUCKETS = [
    (1, "Scheduled depot maintenance & repair",
     "Planned depot-level work to restore or sustain existing ship material condition during a major availability",
     "OMN 1B4B; OPN BA1 L26; USCG O&S",
     "DSRA, SRA, DPMA, PMA, PIA, DPIA, ROH, overhaul, drydock",
     "Hull/HM&E repair, inspections, drydocking, overhauls, corrosion control",
     "Capability-adding mods, class-scale SLEP/MMA, engineering-only support"),
    (2, "Continuous, intermediate & emergent maintenance",
     "Repair outside major depot events: shorter-duration, intermediate, continuous, casualty, voyage, or emergent work",
     "OMN 1B4B (CMMP, Expedited); OMN 1B5B (RMC); USCG O&S",
     "CMAV, emergent, casualty repair, voyage repair, dockside",
     "Continuous maintenance, voyage repairs, repair augmentation, short dockside work",
     "Named class-scale life-extension events and pure modernization"),
    (3, "Modernization & alteration installation",
     "Work that changes configuration or adds capability beyond restore-only maintenance",
     "OPN BA1, BA2, BA4; USCG PC&I",
     "DMP, MODPRD, SHIPALT, ORDALT, tech refresh, AIT",
     "Retrofits, A/B kits, new module installs, combat-system upgrades",
     "Restore-only repair, RDT&E test installations, planning-only engineering"),
    (4, "Major life-cycle events (RCOH, SLEP, MMA)",
     "Multi-year, platform-scale recapitalization or service-life extension separated by scale and funding",
     "SCN (RCOH); OPN BA1 (LCC ESLP); USCG PC&I (ISVS)",
     "RCOH, EOH, ERO, SLEP, MMA, midlife, FRAM",
     "Nuclear refueling overhauls, CG SLEPs, buoy-tender MMA, major midlife events",
     "Ordinary SRA/DSRA/CMAV unless part of a named life-extension program"),
    (5, "Sustainment engineering, planning & obsolescence support",
     "Engineering, planning, lifecycle, DMSMS, and program-support work; not hands-on repair or install",
     "OMN 1B5B; OPN BA8; USCG O&S",
     "SETA, lifecycle support, DMSMS, obsolescence, class maintenance plan, SUPSHIP",
     "Availability planning, technical authority support, configuration management",
     "Physical repair labor, install labor, husbanding/port services"),
    (6, "Availability support, husbanding & port services",
     "Services that enable repair/modernization but are not themselves repair or upgrade work",
     "OMN 1B5B (Berthing & Messing); USCG O&S",
     "Berthing, crane, shore power, scaffolding, force protection, waste removal",
     "Integrated port services, force protection, berthing, cranes, environmental services",
     "Actual ship repair/install work; shipyard infrastructure CAPEX"),
]

# ── Bucket 1 line items ──────────────────────────────────────────────────────
# Tuple: (source, line_item, performer, fy24, fy25, fy26, confidence, notes, exhibit, pdf_page)
B1_LINES = [
    ("OMN 1B4B", "Ship Modernization Program (SMP)", "Mixed", 1203816, 869253, 1197180, "High", "Year-to-year variation reflects availability phasing", "OP-5", pr("OMN_Book.txt", 5596, 5618)),
    ("OMN 1B4B", "NME / 2-Hatch Ships", "Mixed", 1197197, 1306411, 1341155, "High", "", "OP-5", pr("OMN_Book.txt", 5596, 5618)),
    ("OMN 1B4B", "NME / 4-Hatch Ships", "Mixed", 1247653, 1290155, 1289894, "High", "", "OP-5", pr("OMN_Book.txt", 5596, 5618)),
    ("OMN 1B4B", "NME / Submarine Availability", "Public Yard", 871537, 896253, 924629, "High", "Primarily NNSY / PHNSY", "OP-5", pr("OMN_Book.txt", 5596, 5618)),
    ("OMN 1B4B", "NME / Industrial Base", "Mixed", 1245621, 1270256, 1267865, "High", "", "OP-5", pr("OMN_Book.txt", 5596, 5618)),
    ("OMN 1B4B", "NME / NAVPLAN", "Mixed", 1134835, 1139556, 1134328, "High", "", "OP-5", pr("OMN_Book.txt", 5596, 5618)),
    ("OMN 1B4B", "NME / SSN-to-Repair", "Public Yard", 630555, 639229, 642190, "High", "Public yard submarine repair", "OP-5", pr("OMN_Book.txt", 5596, 5618)),
    ("OMN 1B4B", "NME / Carrier", "Mixed", 762048, 816344, 784000, "High", "", "OP-5", pr("OMN_Book.txt", 5596, 5618)),
    ("OMN 1B4B", "NME / Amphibious", "Private", 569696, 656000, 624506, "High", "", "OP-5", pr("OMN_Book.txt", 5596, 5618)),
    ("OMN 1B4B", "NME / LCS", "Private", 481854, 539000, 492493, "High", "", "OP-5", pr("OMN_Book.txt", 5596, 5618)),
    ("OMN 1B4B", "NME / DDG", "Private", 758318, 758610, 749000, "High", "", "OP-5", pr("OMN_Book.txt", 5596, 5618)),
    ("OMN 1B4B", "MSC M&R (transfer from NWCF)", "Private", 0, 0, 1584000, "Medium", "New FY26 transfer; no FY24/FY25 comparables", "OP-5", pr("OMN_Book.txt", 5637, 5651)),
    ("OMN 1B4B", "Residual NME rounding / discrepancy", "Mixed", None, None, -176024, "Low", "SAG total controls; sub-element sum slightly below", "OP-5", pr("OMN_Book.txt", 5423)),
    ("OPN BA1", "L26 Ship Maint, Repair & Modernization", "Private", 2839179, 2392190, 2392620, "High", "", "P-1 / P-40", pr("OPN_BA1_Book.txt", 466, 495)),
    ("USCG O&S", "Bucket 1 share (60% of vessel O&S 25.7)", "Mixed", 329409, 335840, 369358, "Medium", "Analyst allocation: 60% of ~$616M vessel share", "CJ", pr("USCG_Justification.txt", 1410, 1889)),
]

# ── Bucket 2 line items ──────────────────────────────────────────────────────
B2_LINES = [
    ("OMN 1B4B", "CMMP (Continuous Maint & Modernization)", "Mixed", 1259925, 1303227, 1321948, "High", "", "OP-5", pr("OMN_Book.txt", 5609)),
    ("OMN 1B4B", "Expedited M&M", "Mixed", 581796, 602706, 640000, "High", "", "OP-5", pr("OMN_Book.txt", 5596, 5618)),
    ("OMN 1B5B", "RMC", "Mixed", 22311, 25633, 26529, "High", "Regional Maintenance Centers", "OP-5", pr("OMN_Book.txt", 6776)),
    ("USCG O&S", "Bucket 2 share (40% of vessel O&S 25.7)", "Mixed", 219606, 223893, 246238, "Medium", "Analyst allocation: 40% of ~$616M vessel share", "CJ", pr("USCG_Justification.txt", 1410, 1889)),
]

# ── Bucket 3 line items ──────────────────────────────────────────────────────
B3_BA1 = [
    ("OPN BA1", "L1 Surface Power Equipment", "Private", 12855, 20840, 9978, "High", "", "P-1/P-40", pr("OPN_BA1_Book.txt", 276, 2044)),
    ("OPN BA1", "L2 Surface Combatant HM&E", "Private", 92449, 77592, 62004, "High", "", "P-1/P-40", pr("OPN_BA1_Book.txt", 282, 2742)),
    ("OPN BA1", "L4 Sub Periscope, Imaging & Supt Equip", "Mixed", 272778, 290575, 277413, "High", "", "P-1/P-40", pr("OPN_BA1_Book.txt", 294, 6429)),
    ("OPN BA1", "L5 DDG Mod", "Private", 641676, 861066, 997485, "High", "", "P-1/P-40", pr("OPN_BA1_Book.txt", 298, 7982)),
    ("OPN BA1", "L8 LHA/LHD Midlife", "Private", 102334, 81602, 123384, "High", "", "P-1/P-40", pr("OPN_BA1_Book.txt", 326, 11499)),
    ("OPN BA1", "L11 Submarine Support Equipment", "Mixed", 210652, 293766, 383062, "High", "", "P-1/P-40", pr("OPN_BA1_Book.txt", 338)),
    ("OPN BA1", "L12 Virginia Class Support Equipment", "Mixed", 32055, 43565, 52039, "High", "", "P-1/P-40", pr("OPN_BA1_Book.txt", 342)),
    ("OPN BA1", "L13 LCS Class Support Equipment", "Private", 17816, 7318, 2551, "High", "Declining as LCS decommissions", "P-1/P-40", pr("OPN_BA1_Book.txt", 346)),
    ("OPN BA1", "L15 LPD Class Support Equipment", "Private", 80933, 38115, 125542, "High", "", "P-1/P-40", pr("OPN_BA1_Book.txt", 354)),
    ("OPN BA1", "L16 DDG 1000 Class Support Equipment", "Private", 220771, 340668, 115267, "High", "", "P-1/P-40", pr("OPN_BA1_Book.txt", 358)),
    ("OPN BA1", "L17 Strategic Platform Support Equip", "Mixed", 23756, 53931, 53740, "High", "", "P-1/P-40", pr("OPN_BA1_Book.txt", 366)),
    ("OPN BA1", "L21 LCAC", "Private", 10239, 11013, 20321, "High", "", "P-1/P-40", pr("OPN_BA1_Book.txt", 382)),
    ("OPN BA1", "L28 Reactor Components", "Mixed", 389890, 445974, 399603, "High", "", "P-1/P-40", pr("OPN_BA1_Book.txt", 422)),
    ("OPN BA1", "L33 LCS Common Mission Modules", "Private", 49028, 56105, 38880, "High", "", "P-1/P-40", pr("OPN_BA1_Book.txt", 442)),
    ("OPN BA1", "L34 LCS MCM Mission Modules", "Private", 87828, 118247, 91372, "High", "", "P-1/P-40", pr("OPN_BA1_Book.txt", 446)),
    ("OPN BA1", "L36 LCS SUW Mission Modules", "Private", 12094, 11101, 3790, "High", "", "P-1/P-40", pr("OPN_BA1_Book.txt", 454)),
    ("OPN BA1", "L37 LCS In-Service Modernization", "Private", 154561, 188254, 237728, "High", "", "P-1/P-40", pr("OPN_BA1_Book.txt", 458)),
    ("OPN BA1", "L40 LSD Midlife & Modernization", "Private", 3989, 56667, 4079, "High", "", "P-1/P-40", pr("OPN_BA1_Book.txt", 470)),
]
B3_BA2 = [
    ("OPN BA2", "L43 AN/SQQ-89 (100%)", "Private", 133803, 134637, 144425, "High", "100% in-service", "P-1/P-40", pr("OPN_BA2_Book.txt", 508, 2509)),
    ("OPN BA2", "L44 SSN Acoustic Equipment (80%)", "Mixed", 107116, 117310, 398878, "Medium", "80% of total; FY26 ramp", "P-1/P-40", pr("OPN_BA2_Book.txt", 511, 3985)),
    ("OPN BA2", "L46 Sub Acoustic Warfare System (100%)", "Mixed", 41695, 51514, 56482, "High", "100% in-service", "P-1/P-40", pr("OPN_BA2_Book.txt", 542, 5951)),
    ("OPN BA2", "L50 AN/SLQ-32 SEWIP (100%)", "Private", 325996, 182011, 461380, "High", "100% in-service", "P-1/P-40", pr("OPN_BA2_Book.txt", 557, 8477)),
    ("OPN BA2", "L51 Shipboard IW Exploit (100%)", "Mixed", 367452, 362099, 379908, "High", "100% in-service", "P-1/P-40", pr("OPN_BA2_Book.txt", 563, 10450)),
    ("OPN BA2", "L53 CEC (95%)", "Private", 22776, 23696, 25316, "Medium", "95% of total", "P-1/P-40", pr("OPN_BA2_Book.txt", 573, 12421)),
    ("OPN BA2", "L70 CANES (99%)", "Mixed", 416879, 413558, 530408, "Medium", "99% of total", "P-1/P-40", pr("OPN_BA2_Book.txt", 673, 22847)),
    ("OPN BA2", "L72 CANES-Intell (99%)", "Mixed", 38283, 40128, 45813, "Medium", "99% of total", "P-1/P-40", pr("OPN_BA2_Book.txt", 24043, 24068)),
    ("OPN BA2", "L78 In-Service Radars and Sensors (100%)", "Private", 279545, 222607, 249656, "High", "100% in-service", "P-1/P-40", pr("OPN_BA2_Book.txt", 25388)),
    ("OPN BA2", "L79 Battle Force Tactical Network (100%)", "Mixed", 68327, 104119, 106583, "High", "100% in-service", "P-1/P-40", pr("OPN_BA2_Book.txt", 27888)),
    ("OPN BA2", "L80 Shipboard Tactical Comms (90%)", "Mixed", 14440, 15480, 18810, "Medium", "90% of total", "P-1/P-40", pr("OPN_BA2_Book.txt", 27230)),
    ("OPN BA2", "L81 Ship Comms Automation (75%)", "Mixed", 96000, 100800, 121556, "Medium", "75% of total", "P-1/P-40", pr("OPN_BA2_Book.txt", 27248)),
]
B3_BA4 = [
    ("OPN BA4", "BLI 125 Ship Gun Systems (100%)", "Private", 16250, 6416, 7358, "High", "100% in-service", "P-1/P-40", pr("OPN_BA4_Book.txt", 1959, 2121)),
    ("OPN BA4", "BLI 126 Harpoon (100%)", "Mixed", 227, 226, 209, "High", "100% in-service", "P-1/P-40", pr("OPN_BA4_Book.txt", 974, 2216)),
    ("OPN BA4", "BLI 127 Ship Missile Support (100%)", "Mixed", 283972, 376830, 455822, "High", "100% in-service", "P-1/P-40", pr("OPN_BA4_Book.txt", 977, 3780)),
    ("OPN BA4", "BLI 128 Tomahawk (70%)", "Mixed", 48696, 55392, 75396, "Medium", "70% of total", "P-1/P-40", pr("OPN_BA4_Book.txt", 980, 4679)),
    ("OPN BA4", "BLI 130 Strategic Missile (50%)", "Mixed", 193500, 210000, 246500, "Medium", "50% of total", "P-1/P-40", pr("OPN_BA4_Book.txt", 990, 5276)),
    ("OPN BA4", "BLI 131 SSN Combat Control (86%)", "Mixed", 67400, 72536, 88554, "Medium", "86% of total", "P-1/P-40", pr("OPN_BA4_Book.txt", 996, 6835)),
    ("OPN BA4", "BLI 132 ASW Support (100%)", "Mixed", 37301, 25362, 25721, "High", "100% in-service", "P-1/P-40", pr("OPN_BA4_Book.txt", 999)),
    ("OPN BA4", "BLI 134 Directed Energy / ODIN (100%)", "Private", 0, 3817, 2976, "High", "New capability", "P-1/P-40", pr("OPN_BA4_Book.txt", 1027, 8976)),
    ("OPN BA4", "BLI 136 Anti-Ship Missile Decoy (100%)", "Mixed", 51100, 75614, 89129, "High", "100% in-service", "P-1/P-40", pr("OPN_BA4_Book.txt", 1036, 9545)),
]
B3_CG = [
    ("USCG PC&I", "ISSS", "Mixed", 25000, 28000, 30000, "Medium", "In-Service Ship Systems", "CJ", pr("USCG_Justification.txt", 3099, 3556)),
    ("USCG PC&I", "C4ISR", "Mixed", 16000, 16000, 10000, "Medium", "", "CJ", pr("USCG_Justification.txt", 3195, 3383)),
    ("USCG PC&I", "Cyber/EMP (50% vessel share)", "Mixed", 10000, 11000, 12900, "Low", "50% vessel share estimate", "CJ", pr("USCG_Justification.txt", 3196, 3504)),
]

# ── Bucket 4 line items ──────────────────────────────────────────────────────
B4_LINES = [
    ("SCN", "CVN RCOH Full Funding (CVN 75)", "Private (HII-NNS)", 0, 811143, 1779011, "High", "Major ramp in FY26", "P-1/P-40", pr("SCN_Book.txt", 357, 6943)),
    ("SCN", "CVN RCOH Cost-to-Complete (CVN 74)", "Private (HII-NNS)", 42422, 669171, 483100, "High", "Completing CVN 74 RCOH", "P-1/P-40", pr("SCN_Book.txt", 357, 6943)),
    ("SCN", "CVN RCOH Outfitting / Post-Delivery", "Private", 19704, 6660, 12200, "High", "", "P-1/P-40", pr("SCN_Book.txt", 6858, 6884)),
    ("SCN", "LCAC SLEP (ESLEP)", "Private (Walashek)", 15286, 45087, 37390, "High", "", "P-1/P-40", pr("SCN_Book.txt", 16308, 16430)),
    ("SCN", "LCAC SLEP Outfitting", "Private", 728, 0, 493, "High", "", "P-1/P-40", pr("SCN_Book.txt", 16405, 16430)),
    ("OPN BA1", "L9 LCC ESLP", "Private", 10522, 7352, 19276, "High", "", "P-1/P-40", pr("OPN_BA1_Book.txt", 330)),
    ("USCG PC&I", "47-ft MLB SLEP", "CG Yard / Private", 43000, 43000, 45000, "Medium", "", "CJ", pr("USCG_Justification.txt", 2249, 2479)),
    ("USCG PC&I", "270-ft WMEC SLEP", "CG Yard", 46200, 46200, 76000, "Medium", "", "CJ", pr("USCG_Justification.txt", 2249, 2479)),
    ("USCG PC&I", "175-ft WLM MMA", "CG Yard", 17800, 17800, 15000, "Medium", "", "CJ", pr("USCG_Justification.txt", 2249, 2479)),
    ("USCG PC&I", "CGC Healy SLEP", "Mixed", 13000, 13000, 11000, "Medium", "", "CJ", pr("USCG_Justification.txt", 2249, 2479)),
    ("USCG PC&I", "418-ft WMSL MMA (new start)", "CG Yard", 0, 0, 5000, "Low", "New program start FY26", "CJ", pr("USCG_Justification.txt", 2249, 2479)),
    ("USCG PC&I", "Boats RB-M SLEP portion", "Private", 1950, 1950, 9270, "Low", "Ramp-up in FY26", "CJ", pr("USCG_Justification.txt", 2249, 2479)),
]

# ── Bucket 5 line items ──────────────────────────────────────────────────────
B5_NAVY = [
    ("OMN 1B5B", "DDG-1000 Maintenance & Support", "Mixed", 183642, 197600, 209225, "High", "", "OP-5", pr("OMN_Book.txt", 6726, 7234)),
    ("OMN 1B5B", "Mine Countermeasures Ship Support", "Mixed", 14157, 7186, 5573, "High", "Declining with MCM decommissions", "OP-5", pr("OMN_Book.txt", 6726, 7234)),
    ("OMN 1B5B", "PRE/PRL CV/CVN Technical Support", "Mixed", 81444, 81505, 86561, "High", "", "OP-5", pr("OMN_Book.txt", 6726, 7234)),
    ("OMN 1B5B", "Service Craft Support, Boats/Targets Rehab", "Mixed", 7548, 8364, 6842, "High", "", "OP-5", pr("OMN_Book.txt", 6726, 7234)),
    ("OMN 1B5B", "Surface & Amphibious Ship Support", "Mixed", 446317, 463167, 451307, "High", "", "OP-5", pr("OMN_Book.txt", 6726, 7234)),
    ("OMN 1B5B", "Nuclear Propulsion Tech Logistics", "Mixed", 327227, 340900, 355445, "High", "", "OP-5", pr("OMN_Book.txt", 6726, 7234)),
    ("OMN 1B5B", "SUPSHIP Costs", "Public Yard", 255302, 269204, 251345, "High", "Supervisor of Shipbuilding offices", "OP-5", pr("OMN_Book.txt", 6726, 7234)),
    ("OMN 1B5B", "NMP Total (Navy Modernization Process)", "Mixed", 415261, 277773, 410893, "High", "", "OP-5", pr("OMN_Book.txt", 6726, 7234)),
    ("OMN 1B5B", "NMMES", "Mixed", 103028, 104083, 115706, "High", "", "OP-5", pr("OMN_Book.txt", 6726, 7234)),
    ("OMN 1B5B", "Smart Work / TOC", "Mixed", 22363, 34068, 23672, "High", "", "OP-5", pr("OMN_Book.txt", 6726, 7234)),
    ("OMN 1B5B", "CSOSS", "Mixed", 12756, 11062, 14954, "High", "", "OP-5", pr("OMN_Book.txt", 6726, 7234)),
    ("OMN 1B5B", "NAVSEA HQ Civilian Personnel & IT", "Public Yard", 571260, 628655, 579450, "High", "HQ overhead; included as enabler", "OP-5", pr("OMN_Book.txt", 6726, 7234)),
]
B5_OTHER = [
    ("OPN BA8", "BLI 9020 Spares and Repair Parts", "Mixed", 736770, 705144, 883630, "High", "", "P-1/P-5a", pr("OPN_BA5-8_Book.txt", 17364, 17659)),
]
B5_CG = [
    ("USCG O&S", "Survey & Design", "CG Yard", 5000, 5000, 5000, "Low", "Analyst estimate", "CJ", pr("USCG_Justification.txt", 3094, 3236)),
    ("USCG O&S", "PO&M (50% vessel share)", "Mixed", 10500, 10500, 11000, "Low", "50% vessel share estimate", "CJ", pr("USCG_Justification.txt", 3660, 3695)),
]

# ── Bucket 6 line items ──────────────────────────────────────────────────────
B6_LINES = [
    ("OMN 1B5B", "Berthing & Messing", "Mixed", 129771, 135039, 141596, "High", "", "OP-5", pr("OMN_Book.txt", 6726, 7234)),
    ("OMN 1B5B", "Facilities / Supply Ops (incl SIOP)", "Public Yard", 120535, 87573, 81780, "High", "", "OP-5", pr("OMN_Book.txt", 6726, 7234)),
    ("OMN 1B5B", "Other embedded support", "Mixed", 2200, 2200, 2200, "Low", "Estimated", "", ""),
    ("USCG O&S", "Other Equipment & Systems vessel share", "Mixed", 2000, 2100, 2200, "Low", "Analyst estimate", "CJ", pr("USCG_Justification.txt", 1410)),
]

# Roll-up summary
ROLLUP = [
    (1, "Scheduled Depot Maintenance & Repair", 13271718, 12909097, 13617194, "High"),
    (2, "Continuous / Intermediate / Emergent", 2083638, 2155459, 2234715, "High"),
    (3, "Modernization & Alteration Installation", 5505882, 5994559, 6624111, "High-Med"),
    (4, "Major Life-Cycle Events (RCOH/SLEP/MMA)", 210612, 1661363, 2492740, "High"),
    (5, "Sustainment Engineering / Planning", 3192575, 3143211, 3411003, "High"),
    (6, "Availability Support / Port Services", 254506, 226912, 227776, "Medium"),
]
ROLLUP_TOTAL = (24518931, 26090601, 28607539)

# Funding source summary
FUNDING = [
    ("OMN (1B4B + 1B5B)", 17280963, 0.604, "1, 2, 5, 6"),
    ("OPN (BA1 + BA2 + BA4 + BA8)", 8455849, 0.296, "1, 3, 4, 5"),
    ("SCN", 2331470, 0.081, "4"),
    ("USCG O&S", 633796, 0.022, "1, 2, 5, 6"),
    ("USCG PC&I", 203170, 0.007, "3, 4"),
]

# Performer summary
PERFORMERS = [
    ("Public Yard", 2580394, 0.090),
    ("Private Industry", 10923489, 0.382),
    ("Mixed (Public + Private / TBD)", 15103656, 0.528),
]

# Confidence tiers
CONF_TIERS = [
    ("High", 23935495, 0.84),
    ("Medium", 4391368, 0.15),
    ("Low", 227776, 0.01),
]

# Reconciliation checks
RECON_CHECKS = [
    ("OMN 1B4B", "Bucket 1+2 OMN sum vs SAG total", 12817164, 13803188, 986024, "REASONABLE", "Delta from MSC transfer accounting"),
    ("OMN 1B5B", "Bucket 2+5+6 sum vs SAG total", 2761278, 2760878, 400, "PASS", "Rounding only"),
    ("OPN BA1", "Extracted lines vs BA1 total", 5411134, 8021677, 2610543, "EXPECTED", "Delta is new-construction / out-of-scope lines"),
    ("SCN", "In-service vs total SCN", 2312194, 50000000, None, "EXPECTED", "RCOH is ~4-5% of SCN"),
    ("Cross-approp", "No double-counting", None, None, None, "PASS", "OMN, OPN, SCN are separate appropriations"),
]

# Known gaps
GAPS = [
    ("USCG O&S vessel share", "60% allocation is estimate; true figure requires CG-internal data", "+/- $100-200M on USCG Buckets 1 & 2"),
    ("Bucket 6 undercount", "Availability support embedded in Bucket 1 contracts", "True Bucket 6 could be $1-2B; current is a floor"),
    ("MSC granularity", "$1.584B as single line; no breakdown by ship class", "FY24/FY25 comparability issue"),
    ("Public-private split", "Four-shipyard figure includes overhead", "Need contract-level data for precision"),
    ("OPN BA2/BA4 in-service shares", "Estimated percentages (50-99%)", "Bucket 3 could shift by ~$385M"),
    ("FY24/FY25 USCG detail", "Assumed flat year-over-year for ISVS sub-investments", "Actual variation unknown"),
    ("CLS isolation", "CLS distributed across buckets by work type", "Cannot isolate total CLS from budget books"),
]

# Sources
SOURCES = [
    ("S1", "DoD FMR Volume 11B", "DoD Comptroller", "Defines depot maintenance and repairs; funding logic"),
    ("S2", "DoD FMR Volume 2A, Ch. 1", "DoD Comptroller", "O&M vs RDT&E boundary rules"),
    ("S3", "DoD FMR Volume 2B, Ch. 4", "DoD Comptroller", "Procurement modification exhibits P-3a / P-40"),
    ("S4", "FY2026 O-1 O&M Programs", "DoD Comptroller", "Navy O&M budget lines (1B4B, 1B5B)"),
    ("S5", "FY2026 P-1 Procurement Programs", "DoD Comptroller", "OPN line items for modernization"),
    ("S6", "JFMM Volume II", "NAVSEA", "Navy fleet maintenance vocabulary and availability structure"),
    ("S7", "OPNAVINST 4700.7N", "Dept of the Navy", "Ship modernization program adds capability or reliability"),
    ("S8", "USCG ISVS Program Page", "USCG", "Defines ISVS; distinguishes MMA from SLEP"),
    ("S9", "USCG FY2026 Congressional Justification", "USCG / DHS", "Budget anchor for ISVS in PC&I"),
    ("S10", "USCG FY2025 Congressional Justification", "USCG", "ISVS structure and SLEP/MMA details"),
    ("S11", "PSC Manual April 2024", "Acquisition.gov", "PSC definitions, ship repair codes, mod/install codes"),
    ("S12", "NAICS 2022 Manual", "U.S. Census / OMB", "336611 too broad for TAM taxonomy"),
    ("S13", "CNRMC Master Ship Repair Agreement", "NAVSEA", "CMAV vs SRA execution tiers"),
]


# ══════════════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════

wb = openpyxl.Workbook()

# ── 1.0 Divider tab ─────────────────────────────────────────────────────────

ws0 = wb.active
ws0.title = "1.0 Divider"
ws0.sheet_properties.tabColor = TAB_COLOR
sv_list = SheetViewList()
sv_list.sheetView = [SheetView(showGridLines=False, workbookViewId=0)]
ws0.views = sv_list

for c in range(1, 8):
    cell = ws0.cell(row=1, column=c)
    cell.fill = FILL_BLACK
    cell.font = FONT_TITLE
ws0.cell(row=1, column=1).value = "In-Service Vessel TAM — Top-Down Analysis"

for c in range(1, 8):
    cell = ws0.cell(row=2, column=c)
    cell.fill = FILL_WHITE
    cell.font = FONT_GRAY_ITALIC
ws0.cell(row=2, column=1).value = "US Navy & Coast Guard  |  FY26 President's Budget  |  All figures in $K"

for r in range(3, 6):
    for c in range(1, 8):
        ws0.cell(row=r, column=c).fill = FILL_WHITE

set_row_heights(ws0, 5)
set_col_widths(ws0, {"A": 60, "B": 12, "C": 12, "D": 12, "E": 12, "F": 12, "G": 12})


# ── 1.1 Executive summary ───────────────────────────────────────────────────

ws1 = wb.create_sheet("1.1 Executive summary")
COL_COUNT_1 = 8  # A-H
setup_sheet(ws1, "Executive summary", "Headline TAM, six-bucket breakdown, funding and performer splits  |  All figures in $K unless noted", COL_COUNT_1)
set_col_widths(ws1, {"A": 34, "B": 12, "C": 12, "D": 12, "E": 12, "F": 10, "G": 3, "H": 14})

r = 4

# ── KPI box
for c in range(1, COL_COUNT_1 + 1):
    cell = ws1.cell(row=r, column=c)
    cell.fill = FILL_TEAL
ws1.cell(row=r, column=1).value = "Total FY26 TAM"
ws1.cell(row=r, column=1).font = FONT_BOLD
c_kpi = ws1.cell(row=r, column=2)
c_kpi.value = 28607539
c_kpi.font = FONT_BLUE_BOLD
fmt_k(c_kpi, FONT_BLUE_BOLD)
ws1.cell(row=r, column=3).value = "$K"
ws1.cell(row=r, column=3).font = FONT_GRAY
ws1.cell(row=r, column=4).value = "~$28.6B"
ws1.cell(row=r, column=4).font = FONT_BOLD
r += 1
white_row(ws1, r, COL_COUNT_1)
r += 1

# ── Bucket roll-up table
section_header(ws1, r, "TAM by bucket", COL_COUNT_1)
r += 1
set_col_headers(ws1, r, ["Bucket", "FY24 ($K)", "FY25 ($K)", "FY26 ($K)", "% of TAM", "Confidence", "", ""], COL_COUNT_1)
r += 1

for bnum, bname, fy24, fy25, fy26, conf in ROLLUP:
    vals = [f"{bnum}. {bname}", fy24, fy25, fy26, fy26 / ROLLUP_TOTAL[2], conf, None, None]
    fonts = {
        2: FONT_BLUE, 3: FONT_BLUE, 4: FONT_BLUE,
        5: FONT_BLACK_ITALIC,
        6: FONT_GRAY,
    }
    write_data_row(ws1, r, vals, COL_COUNT_1, fonts=fonts)
    fmt_k(ws1.cell(row=r, column=2), FONT_BLUE)
    fmt_k(ws1.cell(row=r, column=3), FONT_BLUE)
    fmt_k(ws1.cell(row=r, column=4), FONT_BLUE)
    fmt_pct(ws1.cell(row=r, column=5))
    r += 1

# Total row
vals = ["Total", ROLLUP_TOTAL[0], ROLLUP_TOTAL[1], ROLLUP_TOTAL[2], 1.0, "", None, None]
write_data_row(ws1, r, vals, COL_COUNT_1, fonts={2: FONT_BOLD, 3: FONT_BOLD, 4: FONT_BOLD, 5: FONT_BLACK_ITALIC}, border=BORDER_MED_TOP)
fmt_k(ws1.cell(row=r, column=2), FONT_BOLD)
fmt_k(ws1.cell(row=r, column=3), FONT_BOLD)
fmt_k(ws1.cell(row=r, column=4), FONT_BOLD)
fmt_pct(ws1.cell(row=r, column=5))
r += 1
white_row(ws1, r, COL_COUNT_1)
r += 1

# ── Funding source
section_header(ws1, r, "By funding source (FY26)", COL_COUNT_1)
r += 1
set_col_headers(ws1, r, ["Appropriation", "FY26 ($K)", "% of TAM", "Buckets served", "", "", "", ""], COL_COUNT_1)
r += 1
for name, amt, pct, buckets in FUNDING:
    write_data_row(ws1, r, [name, amt, pct, buckets, None, None, None, None], COL_COUNT_1,
                   fonts={2: FONT_BLUE, 3: FONT_BLACK_ITALIC, 4: FONT_GRAY})
    fmt_k(ws1.cell(row=r, column=2), FONT_BLUE)
    fmt_pct(ws1.cell(row=r, column=3))
    r += 1
white_row(ws1, r, COL_COUNT_1)
r += 1

# ── Performer
section_header(ws1, r, "By performer (FY26)", COL_COUNT_1)
r += 1
set_col_headers(ws1, r, ["Performer", "FY26 ($K)", "% of TAM", "", "", "", "", ""], COL_COUNT_1)
r += 1
for name, amt, pct in PERFORMERS:
    write_data_row(ws1, r, [name, amt, pct, None, None, None, None, None], COL_COUNT_1,
                   fonts={2: FONT_BLUE, 3: FONT_BLACK_ITALIC})
    fmt_k(ws1.cell(row=r, column=2), FONT_BLUE)
    fmt_pct(ws1.cell(row=r, column=3))
    r += 1
write_data_row(ws1, r, ["Total", 28607539, 1.0, None, None, None, None, None], COL_COUNT_1,
               fonts={2: FONT_BOLD, 3: FONT_BLACK_ITALIC}, border=BORDER_MED_TOP)
fmt_k(ws1.cell(row=r, column=2), FONT_BOLD)
fmt_pct(ws1.cell(row=r, column=3))
r += 1
white_row(ws1, r, COL_COUNT_1)
r += 1

# ── Confidence
section_header(ws1, r, "Confidence summary", COL_COUNT_1)
r += 1
set_col_headers(ws1, r, ["Tier", "FY26 ($K)", "% of TAM", "", "", "", "", ""], COL_COUNT_1)
r += 1
conf_fills = {"High": FILL_GREEN, "Medium": FILL_ORANGE, "Low": FILL_RED}
for tier, amt, pct in CONF_TIERS:
    write_data_row(ws1, r, [tier, amt, pct, None, None, None, None, None], COL_COUNT_1,
                   fonts={2: FONT_BLUE, 3: FONT_BLACK_ITALIC},
                   fills={1: conf_fills.get(tier, FILL_WHITE)})
    fmt_k(ws1.cell(row=r, column=2), FONT_BLUE)
    fmt_pct(ws1.cell(row=r, column=3))
    r += 1
white_row(ws1, r, COL_COUNT_1)
r += 1

# ── Key notes
section_header(ws1, r, "Key notes", COL_COUNT_1)
r += 1
notes = [
    "Market name: In-Service Vessel Sustainment, Repair & Modernization (not MRO)",
    "Private-industry addressable TAM: ~$18-22B (excludes public yard work)",
    "MSC M&R ($1.584B) is FY26 accounting transfer from NWCF; inflates FY24-FY26 growth",
    "Bucket 6 ($228M) is a floor; true availability support is $1-2B embedded in Bucket 1",
    "CVN RCOH phasing drives Bucket 4 FY24-FY26 jump (not a structural trend)",
]
for note in notes:
    write_data_row(ws1, r, [note] + [None] * 7, COL_COUNT_1, fonts={1: FONT_GRAY})
    r += 1

set_row_heights(ws1, r)


# ── 1.2 Taxonomy ─────────────────────────────────────────────────────────────

ws2 = wb.create_sheet("1.2 Taxonomy")
COL_COUNT_2 = 7
setup_sheet(ws2, "Taxonomy", "Six work-type buckets derived from official Navy and Coast Guard definitions, budget structure, and contract language", COL_COUNT_2)
set_col_widths(ws2, {"A": 6, "B": 30, "C": 34, "D": 30, "E": 30, "F": 30, "G": 30})

r = 4
set_col_headers(ws2, r, ["#", "Bucket", "Definition", "Budget anchors", "Include", "Exclude", "Availability keywords"], COL_COUNT_2)
r += 1

for bnum, bname, defn, anchor, keywords, incl, excl in BUCKETS:
    write_data_row(ws2, r, [bnum, bname, defn, anchor, incl, excl, keywords], COL_COUNT_2,
                   fonts={1: FONT_BLUE, 3: FONT_GRAY, 4: FONT_GRAY, 5: FONT_GRAY, 6: FONT_GRAY, 7: FONT_GRAY})
    r += 1

white_row(ws2, r, COL_COUNT_2)
r += 1

# Classification rule
section_header(ws2, r, "Classification rule", COL_COUNT_2)
r += 1
rules = [
    "Restore existing capability / condition  -->  Bucket 1 or 2 (maintenance & repair)",
    "Add capability or change configuration  -->  Bucket 3 (modernization / alteration)",
    "Extend service life at platform scale  -->  Bucket 4 (life-extension / recapitalization)",
    "Plan, engineer, analyze, or manage without doing the repair/install  -->  Bucket 5 (sustainment engineering)",
    "Enable the availability with shore/port/support services  -->  Bucket 6 (availability support)",
]
for rule in rules:
    write_data_row(ws2, r, [None, rule] + [None] * 5, COL_COUNT_2, fonts={2: FONT_GRAY})
    r += 1

set_row_heights(ws2, r)


# ── Helper: write a detail tab with line items ───────────────────────────────

def write_detail_tab(ws, tab_title, purpose, sections, col_count=13):
    """
    sections: list of (section_name, subsections)
    subsections: list of (subsection_name_or_None, lines)
    lines: list of (source, line_item, performer, fy24, fy25, fy26, confidence, notes, exhibit, pdf_page)
    """
    setup_sheet(ws, tab_title, purpose, col_count)
    set_col_widths(ws, {
        "A": 32, "B": 10, "C": 12, "D": 12, "E": 12,
        "F": 3, "G": 10, "H": 12, "I": 10, "J": 12,
        "K": 30, "L": 12, "M": 3
    })

    r = 4
    headers = ["Line item", "Source", "FY24 ($K)", "FY25 ($K)", "FY26 ($K)",
               "", "Confidence", "Performer", "Exhibit", "PDF page",
               "Notes", "Bucket", ""]
    set_col_headers(ws, r, headers, col_count)
    r += 1

    grand_fy24, grand_fy25, grand_fy26 = 0, 0, 0

    for sec_name, subsections in sections:
        # Section header
        section_header(ws, r, sec_name, col_count)
        r += 1

        sec_fy24, sec_fy25, sec_fy26 = 0, 0, 0

        for sub_name, lines in subsections:
            if sub_name:
                subsection_header(ws, r, sub_name, col_count)
                r += 1

            sub_fy24, sub_fy25, sub_fy26 = 0, 0, 0
            for source, item, perf, fy24, fy25, fy26, conf, notes, exhibit, pdf_page in lines:
                vals = [item, source, fy24, fy25, fy26, None, conf, perf, exhibit, pdf_page, notes, None, None]
                fonts_d = {
                    2: FONT_GRAY, 3: FONT_BLUE, 4: FONT_BLUE, 5: FONT_BLUE,
                    7: FONT_GRAY, 8: FONT_GRAY, 9: FONT_GRAY, 10: FONT_GRAY,
                    11: FONT_GRAY, 12: FONT_GRAY,
                }
                conf_fill = {}
                if conf == "Medium":
                    conf_fill = {7: FILL_ORANGE}
                elif conf == "Low":
                    conf_fill = {7: FILL_RED}
                elif conf == "High":
                    conf_fill = {7: FILL_GREEN}

                write_data_row(ws, r, vals, col_count, fonts=fonts_d, fills=conf_fill, indent=1)
                for cc in [3, 4, 5]:
                    fmt_k(ws.cell(row=r, column=cc), FONT_BLUE)
                r += 1

                if fy24 is not None:
                    sub_fy24 += fy24
                if fy25 is not None:
                    sub_fy25 += fy25
                if fy26 is not None:
                    sub_fy26 += fy26

            if sub_name and len(lines) > 1:
                # Subsection subtotal
                vals = [f"  {sub_name} subtotal", "", sub_fy24, sub_fy25, sub_fy26] + [None] * (col_count - 5)
                write_data_row(ws, r, vals, col_count,
                               fonts={1: FONT_BOLD, 3: FONT_BOLD, 4: FONT_BOLD, 5: FONT_BOLD},
                               border=BORDER_THIN_TOP)
                for cc in [3, 4, 5]:
                    fmt_k(ws.cell(row=r, column=cc), FONT_BOLD)
                r += 1

            sec_fy24 += sub_fy24
            sec_fy25 += sub_fy25
            sec_fy26 += sub_fy26

        # Section total
        vals = [f"{sec_name} total", "", sec_fy24, sec_fy25, sec_fy26] + [None] * (col_count - 5)
        write_data_row(ws, r, vals, col_count,
                       fonts={1: FONT_BOLD, 3: FONT_BOLD, 4: FONT_BOLD, 5: FONT_BOLD},
                       border=BORDER_MED_TOP)
        for cc in [3, 4, 5]:
            fmt_k(ws.cell(row=r, column=cc), FONT_BOLD)
        r += 1
        white_row(ws, r, col_count)
        r += 1

        grand_fy24 += sec_fy24
        grand_fy25 += sec_fy25
        grand_fy26 += sec_fy26

    # Grand total
    vals = ["Grand total", "", grand_fy24, grand_fy25, grand_fy26] + [None] * (col_count - 5)
    write_data_row(ws, r, vals, col_count,
                   fonts={1: FONT_BOLD, 3: FONT_BOLD, 4: FONT_BOLD, 5: FONT_BOLD},
                   border=BORDER_THICK_TOP)
    for cc in [3, 4, 5]:
        fmt_k(ws.cell(row=r, column=cc), FONT_BOLD)
    r += 1

    set_row_heights(ws, r)
    return r


# ── 1.3 Navy top-down ────────────────────────────────────────────────────────

ws3 = wb.create_sheet("1.3 Navy top-down")

navy_sections = [
    ("Bucket 1: Scheduled Depot Maintenance & Repair", [
        ("OMN 1B4B", [l for l in B1_LINES if l[0].startswith("OMN")]),
        ("OPN BA1", [l for l in B1_LINES if l[0].startswith("OPN")]),
    ]),
    ("Bucket 2: Continuous / Intermediate / Emergent", [
        (None, [l for l in B2_LINES if not l[0].startswith("USCG")]),
    ]),
    ("Bucket 3: Modernization & Alteration Installation", [
        ("OPN BA1 — Ships Support Equipment", B3_BA1),
        ("OPN BA2 — Communications & Electronics (adjusted)", B3_BA2),
        ("OPN BA4 — Ordnance Support (adjusted)", B3_BA4),
    ]),
    ("Bucket 4: Major Life-Cycle Events", [
        (None, [l for l in B4_LINES if not l[0].startswith("USCG")]),
    ]),
    ("Bucket 5: Sustainment Engineering / Planning", [
        ("OMN 1B5B", B5_NAVY),
        ("OPN BA8", B5_OTHER),
    ]),
    ("Bucket 6: Availability Support / Port Services", [
        (None, [l for l in B6_LINES if not l[0].startswith("USCG")]),
    ]),
]

write_detail_tab(ws3, "Navy top-down detail",
                 "All Navy line items from PB26 justification books (OMN, OPN, SCN)  |  $K",
                 navy_sections)


# ── 1.4 Coast Guard top-down ─────────────────────────────────────────────────

ws4 = wb.create_sheet("1.4 Coast Guard top-down")

cg_sections = [
    ("Bucket 1: Scheduled Depot Maintenance & Repair", [
        (None, [l for l in B1_LINES if l[0].startswith("USCG")]),
    ]),
    ("Bucket 2: Continuous / Intermediate / Emergent", [
        (None, [l for l in B2_LINES if l[0].startswith("USCG")]),
    ]),
    ("Bucket 3: Modernization & Alteration Installation", [
        (None, B3_CG),
    ]),
    ("Bucket 4: Major Life-Cycle Events (SLEP / MMA)", [
        (None, [l for l in B4_LINES if l[0].startswith("USCG")]),
    ]),
    ("Bucket 5: Sustainment Engineering / Planning", [
        (None, B5_CG),
    ]),
    ("Bucket 6: Availability Support / Port Services", [
        (None, [l for l in B6_LINES if l[0].startswith("USCG")]),
    ]),
]

write_detail_tab(ws4, "Coast Guard top-down detail",
                 "All USCG line items from FY26 Congressional Justification  |  $K",
                 cg_sections)


# ── 1.5 Consolidated TAM ─────────────────────────────────────────────────────

ws5 = wb.create_sheet("1.5 Consolidated TAM")
COL_COUNT_5 = 9
setup_sheet(ws5, "Consolidated TAM", "Six-bucket roll-up: Navy + Coast Guard = Total  |  $K", COL_COUNT_5)
set_col_widths(ws5, {"A": 34, "B": 12, "C": 12, "D": 12, "E": 3, "F": 12, "G": 12, "H": 12, "I": 12})

r = 4
set_col_headers(ws5, r, ["Bucket", "Navy FY26", "CG FY26", "Total FY26", "", "FY24", "FY25", "FY26", "% of TAM"], COL_COUNT_5)
r += 1

# Navy/CG splits from silver canonical
navy_cg_splits = [
    (1, "Scheduled Depot Maintenance & Repair", 13247836, 369358, 13617194, 13271718, 12909097, 13617194),
    (2, "Continuous / Intermediate / Emergent", 1988477, 246238, 2234715, 2083638, 2155459, 2234715),
    (3, "Modernization & Alteration Installation", 6571211, 52900, 6624111, 5505882, 5994559, 6624111),
    (4, "Major Life-Cycle Events", 2331470, 161270, 2492740, 210612, 1661363, 2492740),
    (5, "Sustainment Engineering / Planning", 3395003, 16000, 3411003, 3192575, 3143211, 3411003),
    (6, "Availability Support / Port Services", 225576, 2200, 227776, 254506, 226912, 227776),
]

for bnum, bname, navy, cg, total, fy24, fy25, fy26 in navy_cg_splits:
    pct = fy26 / ROLLUP_TOTAL[2]
    vals = [f"{bnum}. {bname}", navy, cg, total, None, fy24, fy25, fy26, pct]
    fonts_d = {
        2: FONT_BLUE, 3: FONT_BLUE, 4: FONT_BOLD,
        6: FONT_BLUE, 7: FONT_BLUE, 8: FONT_BLUE, 9: FONT_BLACK_ITALIC,
    }
    write_data_row(ws5, r, vals, COL_COUNT_5, fonts=fonts_d)
    for cc in [2, 3, 4, 6, 7, 8]:
        fmt_k(ws5.cell(row=r, column=cc))
    fmt_pct(ws5.cell(row=r, column=9))
    r += 1

# Total
navy_total = sum(x[2] for x in navy_cg_splits)
cg_total = sum(x[3] for x in navy_cg_splits)
vals = ["Total", navy_total, cg_total, ROLLUP_TOTAL[2], None, ROLLUP_TOTAL[0], ROLLUP_TOTAL[1], ROLLUP_TOTAL[2], 1.0]
write_data_row(ws5, r, vals, COL_COUNT_5,
               fonts={1: FONT_BOLD, 2: FONT_BOLD, 3: FONT_BOLD, 4: FONT_BOLD,
                      6: FONT_BOLD, 7: FONT_BOLD, 8: FONT_BOLD, 9: FONT_BLACK_ITALIC},
               border=BORDER_MED_TOP)
for cc in [2, 3, 4, 6, 7, 8]:
    fmt_k(ws5.cell(row=r, column=cc), FONT_BOLD)
fmt_pct(ws5.cell(row=r, column=9))
r += 1
white_row(ws5, r, COL_COUNT_5)
r += 1

# Customer split
section_header(ws5, r, "Customer split (FY26)", COL_COUNT_5)
r += 1
write_data_row(ws5, r, ["Navy", navy_total, None, None, None, None, None, None, navy_total / ROLLUP_TOTAL[2]], COL_COUNT_5,
               fonts={2: FONT_BLUE, 9: FONT_BLACK_ITALIC})
fmt_k(ws5.cell(row=r, column=2), FONT_BLUE)
fmt_pct(ws5.cell(row=r, column=9))
r += 1
write_data_row(ws5, r, ["Coast Guard", cg_total, None, None, None, None, None, None, cg_total / ROLLUP_TOTAL[2]], COL_COUNT_5,
               fonts={2: FONT_BLUE, 9: FONT_BLACK_ITALIC})
fmt_k(ws5.cell(row=r, column=2), FONT_BLUE)
fmt_pct(ws5.cell(row=r, column=9))
r += 1

set_row_heights(ws5, r)


# ── 1.6 Confidence and gaps ──────────────────────────────────────────────────

ws6 = wb.create_sheet("1.6 Confidence and gaps")
COL_COUNT_6 = 7
setup_sheet(ws6, "Confidence and gaps", "Confidence tiers, known data gaps, sensitivity analysis, and known exclusions", COL_COUNT_6)
set_col_widths(ws6, {"A": 30, "B": 14, "C": 12, "D": 30, "E": 30, "F": 30, "G": 3})

r = 4

# Confidence tiers
section_header(ws6, r, "Confidence tiers (FY26)", COL_COUNT_6)
r += 1
set_col_headers(ws6, r, ["Tier", "Amount ($K)", "% of TAM", "", "", "", ""], COL_COUNT_6)
r += 1
for tier, amt, pct in CONF_TIERS:
    cf = conf_fills.get(tier, FILL_WHITE)
    write_data_row(ws6, r, [tier, amt, pct, None, None, None, None], COL_COUNT_6,
                   fonts={2: FONT_BLUE, 3: FONT_BLACK_ITALIC}, fills={1: cf})
    fmt_k(ws6.cell(row=r, column=2), FONT_BLUE)
    fmt_pct(ws6.cell(row=r, column=3))
    r += 1
white_row(ws6, r, COL_COUNT_6)
r += 1

# Known gaps
section_header(ws6, r, "Known data gaps", COL_COUNT_6)
r += 1
set_col_headers(ws6, r, ["Gap", "Description", "", "Estimated impact", "", "", ""], COL_COUNT_6)
r += 1
for gap_name, desc, impact in GAPS:
    write_data_row(ws6, r, [gap_name, desc, None, impact, None, None, None], COL_COUNT_6,
                   fonts={1: FONT_BASE, 2: FONT_GRAY, 4: FONT_GRAY},
                   fills={1: FILL_ORANGE})
    r += 1
white_row(ws6, r, COL_COUNT_6)
r += 1

# Sensitivity
section_header(ws6, r, "Sensitivity analysis", COL_COUNT_6)
r += 1
set_col_headers(ws6, r, ["Scenario", "Effect on TAM", "", "", "", "", ""], COL_COUNT_6)
r += 1
sensitivities = [
    ("USCG O&S vessel share 50% instead of 60%", "TAM decreases by ~$103M"),
    ("USCG O&S vessel share 70% instead of 60%", "TAM increases by ~$103M"),
    ("Bucket 6 embedded costs fully separated", "Bucket 6 +$1-2B; Bucket 1 same decrease; net TAM unchanged"),
    ("OPN BA2/BA4 in-service shares all 100%", "Bucket 3 increases by ~$385M"),
    ("MSC M&R excluded (FY26 transfer)", "Bucket 1 decreases by $1,584,000K"),
]
for scenario, effect in sensitivities:
    write_data_row(ws6, r, [scenario, effect, None, None, None, None, None], COL_COUNT_6,
                   fonts={2: FONT_GRAY})
    r += 1
white_row(ws6, r, COL_COUNT_6)
r += 1

# Reconciliation checks
section_header(ws6, r, "Internal consistency checks", COL_COUNT_6)
r += 1
set_col_headers(ws6, r, ["Check", "Our sum ($K)", "Control ($K)", "Delta ($K)", "Status", "Explanation", ""], COL_COUNT_6)
r += 1
for check_name, desc, our_sum, control, delta, status, expl in RECON_CHECKS:
    status_fill = FILL_GREEN if status == "PASS" else (FILL_ORANGE if status in ("REASONABLE", "EXPECTED") else FILL_RED)
    vals = [f"{check_name}: {desc}", our_sum, control, delta, status, expl, None]
    write_data_row(ws6, r, vals, COL_COUNT_6,
                   fonts={2: FONT_BLUE, 3: FONT_BLUE, 4: FONT_BLUE, 5: FONT_BOLD, 6: FONT_GRAY},
                   fills={5: status_fill})
    for cc in [2, 3, 4]:
        if ws6.cell(row=r, column=cc).value is not None:
            fmt_k(ws6.cell(row=r, column=cc), FONT_BLUE)
    r += 1

set_row_heights(ws6, r)


# ── 1.7 Methodology and sources ──────────────────────────────────────────────

ws7 = wb.create_sheet("1.7 Methodology and sources")
COL_COUNT_7 = 6
setup_sheet(ws7, "Methodology and sources", "Hierarchy of evidence, key assumptions, and source index", COL_COUNT_7)
set_col_widths(ws7, {"A": 6, "B": 28, "C": 30, "D": 30, "E": 30, "F": 14})

r = 4

# Hierarchy of evidence
section_header(ws7, r, "Hierarchy of evidence", COL_COUNT_7)
r += 1
set_col_headers(ws7, r, ["Rank", "Evidence type", "Use it for", "Why it works", "Main weakness", "Sources"], COL_COUNT_7)
r += 1
evidence = [
    (1, "Budget / program line / appropriation", "Top-level TAM taxonomy + top-down sizing", "Matches how the customer funds and governs work", "Less retrieval-friendly for contract pulls", "S1-S5, S8-S10"),
    (2, "Official work-type definitions", "Boundary rules between maint, mod, life ext", "Gives defensible logic for what counts where", "Services phrase work differently", "S1-S3, S6-S8"),
    (3, "PSC (predominant product/service)", "Primary bottom-up coding spine", "Designed to say what was bought", "Mixed awards collapse to single code", "S11"),
    (4, "Availability acronym + keywords", "Second-level tagging and retrieval", "How ship repair is named in contract text", "Execution construct, not complete taxonomy", "S6, S8, S13"),
    (5, "NAICS", "Broad universe filter only", "Useful for finding shipyards", "Too broad: mixes construction and repair", "S12"),
]
for rank, etype, use, why, weakness, sources in evidence:
    write_data_row(ws7, r, [rank, etype, use, why, weakness, sources], COL_COUNT_7,
                   fonts={1: FONT_BLUE, 3: FONT_GRAY, 4: FONT_GRAY, 5: FONT_GRAY, 6: FONT_GRAY})
    r += 1
white_row(ws7, r, COL_COUNT_7)
r += 1

# Key assumptions
section_header(ws7, r, "Key assumptions", COL_COUNT_7)
r += 1
assumptions = [
    "FY26 President's Budget request is primary TAM year",
    "Aircraft and aviation systems excluded from vessel TAM",
    "New construction excluded; only in-service / post-delivery work",
    "USCG O&S vessel share estimated at 60% of Object Class 25.7",
    "In-service share percentages applied to partial OPN lines (80-99%)",
    "All figures in then-year dollars; not inflation-adjusted",
    "MSC M&R counted once in OMN 1B4B after FY26 transfer from NWCF",
    "NWCF and OMN Vol2 data tagged as performer reference, not counted in TAM",
]
for assumption in assumptions:
    write_data_row(ws7, r, [None, assumption, None, None, None, None], COL_COUNT_7,
                   fonts={2: FONT_GRAY})
    r += 1
white_row(ws7, r, COL_COUNT_7)
r += 1

# Source index
section_header(ws7, r, "Source index", COL_COUNT_7)
r += 1
set_col_headers(ws7, r, ["ID", "Document", "Organization", "Why it matters", "", ""], COL_COUNT_7)
r += 1
for sid, doc, org, why in SOURCES:
    write_data_row(ws7, r, [sid, doc, org, why, None, None], COL_COUNT_7,
                   fonts={1: FONT_BLUE, 3: FONT_GRAY, 4: FONT_GRAY})
    r += 1

set_row_heights(ws7, r)


# ── Save ─────────────────────────────────────────────────────────────────────

OUTPUT_PATH = "topdown_tam_workbook_v2.xlsx"
wb.save(OUTPUT_PATH)
print(f"Workbook saved to {OUTPUT_PATH}")
print(f"Tabs: {[ws.title for ws in wb.worksheets]}")

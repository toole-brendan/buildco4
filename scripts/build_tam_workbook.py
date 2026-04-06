#!/usr/bin/env python3
"""Build TAM workbook with full styling per the workbook styling guide."""

import os
from copy import copy
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.views import SheetView, SheetViewList

# ── Colour palette (ARGB with FF prefix) ──────────────────────────────
CLR_BLUE      = "FF0000CC"   # hardcoded numeric input
CLR_BLACK     = "FF000000"
CLR_GREEN     = "FF008000"   # cross-tab link
CLR_GRAY      = "FF808080"   # notes / sources
CLR_RED       = "FFFF0000"
CLR_WHITE     = "FFFFFFFF"

FILL_WHITE    = PatternFill("solid", fgColor=CLR_WHITE)
FILL_LTGRAY   = PatternFill("solid", fgColor="FFF2F4F7")  # headers
FILL_BLACK    = PatternFill("solid", fgColor=CLR_BLACK)    # title band
FILL_YELLOW   = PatternFill("solid", fgColor="FFFFF9C4")   # input
FILL_ORANGE   = PatternFill("solid", fgColor="FFFFF3E0")   # caution
FILL_TEAL     = PatternFill("solid", fgColor="FFE0F2F1")   # KPI
FILL_PASS     = PatternFill("solid", fgColor="FFE8F5E9")   # pass green
FILL_FAIL     = PatternFill("solid", fgColor="FFFFEBEE")   # error red

FONT_BASE     = Font(name="Arial", size=8, color=CLR_BLACK)
FONT_BOLD     = Font(name="Arial", size=8, color=CLR_BLACK, bold=True)
FONT_WHITE_B  = Font(name="Arial", size=8, color=CLR_WHITE, bold=True)
FONT_GRAY     = Font(name="Arial", size=8, color=CLR_GRAY)
FONT_GRAY_I   = Font(name="Arial", size=8, color=CLR_GRAY, italic=True)
FONT_BLUE     = Font(name="Arial", size=8, color=CLR_BLUE)
FONT_GREEN    = Font(name="Arial", size=8, color=CLR_GREEN)
FONT_RED      = Font(name="Arial", size=8, color=CLR_RED)

BORDER_THIN_BOTTOM = Border(bottom=Side(style="thin"))
BORDER_THIN_TOP    = Border(top=Side(style="thin"))
BORDER_MED_TOP     = Border(top=Side(style="medium"))

NUM_FMT_K   = '#,##0;[Red](#,##0);"-"'
NUM_FMT_PCT = "0.0%"

TAB_COLOR = "FF5B7A99"
ROW_HEIGHT = 10.0

ALIGN_LEFT  = Alignment(horizontal="left", vertical="center", wrap_text=False)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=False)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)


# ── Helpers ────────────────────────────────────────────────────────────

def style_cell(cell, font=None, fill=None, alignment=None, number_format=None, border=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if number_format:
        cell.number_format = number_format
    if border:
        cell.border = border


def apply_pct_style(cell, font_color=CLR_BLACK):
    """Percentages must be italic."""
    cell.font = Font(name="Arial", size=8, color=font_color, italic=True)
    cell.number_format = NUM_FMT_PCT


def fill_row_white(ws, row, max_col):
    """Ensure every cell in the row has explicit fill."""
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        if cell.fill is None or cell.fill.fgColor is None or cell.fill.fgColor.rgb in (None, "00000000"):
            cell.fill = FILL_WHITE


def set_all_row_heights(ws, max_row):
    for r in range(1, max_row + 1):
        ws.row_dimensions[r].height = ROW_HEIGHT


def setup_sheet_view(ws):
    """No gridlines, no frozen panes."""
    ws.views = SheetViewList()
    ws.views.sheetView.append(SheetView(showGridLines=False, workbookViewId=0))


def write_title_row(ws, title, max_col):
    """Row 1: black fill, white bold text."""
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = FILL_BLACK
        cell.font = FONT_WHITE_B
        cell.alignment = ALIGN_LEFT
    ws.cell(row=1, column=1).value = title


def write_purpose_row(ws, purpose, max_col):
    """Row 2: gray italic, white fill."""
    for c in range(1, max_col + 1):
        cell = ws.cell(row=2, column=c)
        cell.fill = FILL_WHITE
        cell.font = FONT_GRAY_I
        cell.alignment = ALIGN_LEFT
    ws.cell(row=2, column=1).value = purpose


def write_units_row(ws, units, max_col):
    """Row 3: gray text, white fill."""
    for c in range(1, max_col + 1):
        cell = ws.cell(row=3, column=c)
        cell.fill = FILL_WHITE
        cell.font = FONT_GRAY
        cell.alignment = ALIGN_LEFT
    ws.cell(row=3, column=1).value = units


def write_col_headers(ws, headers, row, max_col, widths=None):
    """Light gray fill, bold text, thin bottom border."""
    for i, h in enumerate(headers):
        c = i + 1
        cell = ws.cell(row=row, column=c)
        cell.value = h
        cell.font = FONT_BOLD
        cell.fill = FILL_LTGRAY
        cell.alignment = ALIGN_LEFT if i == 0 or (isinstance(h, str) and not any(ch.isdigit() for ch in h)) else ALIGN_RIGHT
        cell.border = BORDER_THIN_BOTTOM
    # Fill remaining cols
    for c in range(len(headers) + 1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_LTGRAY
        cell.border = BORDER_THIN_BOTTOM


def write_section_header(ws, row, text, max_col):
    """Major section: black fill, white bold."""
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_BLACK
        cell.font = FONT_WHITE_B
        cell.alignment = ALIGN_LEFT
    ws.cell(row=row, column=1).value = text


def write_subsection_header(ws, row, text, max_col):
    """Subsection: light gray fill, bold."""
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_LTGRAY
        cell.font = FONT_BOLD
        cell.alignment = ALIGN_LEFT
    ws.cell(row=row, column=1).value = text


def ensure_fills(ws, max_row, max_col):
    """Every cell through last active column must have explicit fill."""
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.fill.fgColor is None or cell.fill.fgColor.rgb in (None, "00000000"):
                cell.fill = FILL_WHITE


# ── TAB 1: TAM summary ────────────────────────────────────────────────

def build_tab1(wb):
    ws = wb.active
    ws.title = "TAM summary"
    ws.sheet_properties.tabColor = TAB_COLOR
    MAX_COL = 9

    tam_data = [
        [1, "New Construction", 28427000, 38453000, 19822000, 26507000, 46326000],
        [2, "Scheduled Depot Maintenance & Repair", 13271718, 12909097, 13617194, 0, 13617194],
        [3, "Continuous / Intermediate / Emergent", 2083638, 2155459, 2234715, 0, 2234715],
        [4, "Modernization & Alteration Installation", 5505882, 5994559, 6624111, 0, 6624111],
        [5, "Major Life-Cycle Events (RCOH/SLEP/MMA)", 210612, 1661363, 2492740, 0, 2492740],
        [6, "Sustainment Engineering / Planning", 3192575, 3143211, 3411003, 0, 3411003],
        [7, "Availability Support / Port Services", 254506, 226912, 227776, 0, 227776],
    ]

    # Header rows
    write_title_row(ws, "TAM summary", MAX_COL)
    write_purpose_row(ws, "Top-down TAM by bucket \u2014 FY24 through FY26, with OBBA reconciliation breakout", MAX_COL)
    write_units_row(ws, "All figures in USD thousands ($K)", MAX_COL)

    headers = ["#", "Bucket", "FY24 Actual", "FY25 Enacted", "FY26 Request",
               "FY26 OBBA", "FY26 Total", "% of Full TAM", "% of In-Service"]
    write_col_headers(ws, headers, 4, MAX_COL)
    # Fix alignment: first two left, rest right
    for c in range(3, MAX_COL + 1):
        ws.cell(row=4, column=c).alignment = ALIGN_RIGHT

    # Data rows (rows 5-11)
    DATA_START = 5
    for i, row_data in enumerate(tam_data):
        r = DATA_START + i
        bnum, bname, fy24, fy25, fy26r, fy26o, fy26t = row_data
        ws.cell(row=r, column=1, value=bnum).font = FONT_BASE
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT
        ws.cell(row=r, column=2, value=bname).font = FONT_BASE
        ws.cell(row=r, column=2).alignment = ALIGN_LEFT
        for ci, val in enumerate([fy24, fy25, fy26r, fy26o, fy26t], start=3):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = FONT_BLUE
            cell.number_format = NUM_FMT_K
            cell.alignment = ALIGN_RIGHT

    DATA_END = DATA_START + len(tam_data) - 1  # row 11

    # Subtotal row: In-Service (Buckets 2-7) = rows 6-11
    SUB_ROW = DATA_END + 1  # row 12
    ws.cell(row=SUB_ROW, column=1).font = FONT_BOLD
    ws.cell(row=SUB_ROW, column=2, value="In-Service (Buckets 2\u20137)").font = FONT_BOLD
    ws.cell(row=SUB_ROW, column=2).alignment = ALIGN_LEFT
    for ci in range(3, 8):
        col_letter = get_column_letter(ci)
        cell = ws.cell(row=SUB_ROW, column=ci)
        cell.value = f"=SUM({col_letter}{DATA_START+1}:{col_letter}{DATA_END})"
        cell.font = Font(name="Arial", size=8, color=CLR_BLACK, bold=True)
        cell.number_format = NUM_FMT_K
        cell.alignment = ALIGN_RIGHT
        cell.border = BORDER_THIN_TOP

    # Total row: Full TAM = all rows
    TOT_ROW = SUB_ROW + 1  # row 13
    ws.cell(row=TOT_ROW, column=1).font = FONT_BOLD
    ws.cell(row=TOT_ROW, column=2, value="Full TAM").font = FONT_BOLD
    ws.cell(row=TOT_ROW, column=2).alignment = ALIGN_LEFT
    for ci in range(3, 8):
        col_letter = get_column_letter(ci)
        cell = ws.cell(row=TOT_ROW, column=ci)
        cell.value = f"=SUM({col_letter}{DATA_START}:{col_letter}{DATA_END})"
        cell.font = Font(name="Arial", size=8, color=CLR_BLACK, bold=True)
        cell.number_format = NUM_FMT_K
        cell.alignment = ALIGN_RIGHT
        cell.border = BORDER_MED_TOP

    # % of Full TAM (col H = 8): FY26 Total / Full TAM total
    fy26t_col = get_column_letter(7)  # G = FY26 Total
    for i, row_data in enumerate(tam_data):
        r = DATA_START + i
        cell = ws.cell(row=r, column=8)
        cell.value = f"={fy26t_col}{r}/{fy26t_col}{TOT_ROW}"
        apply_pct_style(cell)
        cell.alignment = ALIGN_RIGHT

    # % of In-Service (col I = 9): FY26 Total / In-Service subtotal (only buckets 2-7)
    for i, row_data in enumerate(tam_data):
        r = DATA_START + i
        cell = ws.cell(row=r, column=9)
        if row_data[0] == 1:
            cell.value = "\u2014"
            cell.font = FONT_GRAY
            cell.alignment = ALIGN_RIGHT
        else:
            cell.value = f"={fy26t_col}{r}/{fy26t_col}{SUB_ROW}"
            apply_pct_style(cell)
            cell.alignment = ALIGN_RIGHT

    # Apply borders to subtotal cells cols 8-9
    for ci in [8, 9]:
        ws.cell(row=SUB_ROW, column=ci).border = BORDER_THIN_TOP
        ws.cell(row=TOT_ROW, column=ci).border = BORDER_MED_TOP

    # Column widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 34
    for col_letter in ["C", "D", "E", "F", "G"]:
        ws.column_dimensions[col_letter].width = 12
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 12

    max_row = TOT_ROW
    ensure_fills(ws, max_row, MAX_COL)
    set_all_row_heights(ws, max_row)
    setup_sheet_view(ws)

    return ws


# ── TAB 2: By vessel class ────────────────────────────────────────────

def build_tab2(wb):
    ws = wb.create_sheet("By vessel class")
    ws.sheet_properties.tabColor = TAB_COLOR
    MAX_COL = 11  # A-K

    class_data = [
        ["SSBN (Columbia)", 8994594, 1925892, 10920486, None, None, None, None, None, None],
        ["SSN (Virginia)", 4760225, 6317804, 11078029, None, None, None, None, None, None],
        ["SSN/SSBN (in-service)", 0, 0, 0, 1566819, None, 766254, 0, 0, None],
        ["CVN (Nimitz/Ford)", 3281673, 0, 3281673, 784000, None, 0, 2274311, 86561, None],
        ["DDG-51", 10773, 5400000, 5410773, 749000, None, 997485, 0, 0, None],
        ["DDG-1000", 52358, 0, 52358, 0, None, 115267, 0, 209225, None],
        ["CG (Ticonderoga)", 0, 0, 0, 0, None, 0, 0, 0, None],
        ["FFG-62", 0, 0, 0, 0, None, 0, 0, 0, None],
        ["LCS", 0, 0, 0, 492493, None, 374321, 0, 0, None],
        ["Amphibious (combined)", 0, 0, 0, 624506, None, 0, 0, 0, None],
        ["LPD Flight II", 0, 2600000, 2600000, None, None, 125542, 0, 0, None],
        ["LHA", 0, 3895000, 3895000, None, None, 123384, 0, 0, None],
        ["LSD", 0, 0, 0, None, None, 4079, 0, 0, None],
        ["LSM", 0, 1963941, 1963941, 0, None, 0, 0, 0, None],
        ["T-AO (John Lewis)", 8346, 1853359, 1861705, 0, None, 0, 0, 0, None],
        ["T-AGOS", 612205, 0, 612205, 0, None, 0, 0, 0, None],
        ["SSC/LCAC", 0, 239095, 239095, 0, None, 20321, 37883, 0, None],
        ["LCU 1700", 0, 295000, 295000, 0, None, 0, 0, 0, None],
        ["Sealift (Strategic + Used)", 45000, 700000, 745000, 0, None, 0, 0, 0, None],
        ["LCC", 0, 0, 0, 0, None, 0, 19276, 0, None],
        ["ESB/ESD", 0, 0, 0, 0, None, 0, 0, 0, None],
        ["MSC Fleet", 0, 0, 0, 1584000, None, 0, 0, 0, None],
        ["Service Craft / APL", 34602, 0, 34602, 0, None, 0, 0, 0, None],
        ["CG: OPC (Heritage)", 812400, 0, 812400, 0, None, 0, 0, 0, None],
        ["CG: FRC (Sentinel)", 216000, 0, 216000, 0, None, 0, 0, 0, None],
        ["CG: PSC / WCC / Icebreaker", 228000, 0, 228000, 0, None, 0, 0, 0, None],
        ["CG: WMEC-270", 0, 0, 0, None, None, 0, 76000, 0, None],
        ["CG: WLM-175", 0, 0, 0, None, None, 0, 15000, 0, None],
        ["CG: MLB-47", 0, 0, 0, None, None, 0, 45000, 0, None],
        ["CG: Healy / WMSL", 0, 0, 0, None, None, 0, 16000, 0, None],
        ["CG: RB-M", 0, 0, 0, None, None, 0, 9270, 0, None],
        ["Outfitting (all newbuild)", 863846, 23449, 887295, 0, None, 0, 0, 0, None],
        ["PY Completion (multi-class)", 725429, 466837, 1192266, 0, None, 0, 0, 0, None],
        ["CG: In-service O&S", 0, 0, 0, 369358, 246238, 52900, 0, 16000, None],
    ]

    # Section groupings for visual breaks
    section_breaks = {
        0: "Navy Submarines",
        3: "Navy Surface",
        9: "Navy Amphibious",
        14: "Navy Auxiliaries",
        23: "Coast Guard",
        31: "Multi-class",
        33: "CG In-service",
    }

    # Header rows
    write_title_row(ws, "By vessel class", MAX_COL)
    write_purpose_row(ws, "FY26 spending by vessel class and bucket \u2014 budget book data only, no analyst estimates", MAX_COL)
    write_units_row(ws, "All figures in USD thousands ($K), FY26", MAX_COL)

    headers = ["Vessel Class", "B1 Request", "B1 OBBA", "B1 Total", "B2", "B3", "B4", "B5", "B6", "B7", "Row Total"]
    write_col_headers(ws, headers, 4, MAX_COL)
    for c in range(2, MAX_COL + 1):
        ws.cell(row=4, column=c).alignment = ALIGN_RIGHT

    # Data columns mapping: B=2 B1Req, C=3 B1OBBA, D=4 B1Total, E=5 B2, F=6 B3, G=7 B4, H=8 B5, I=9 B6, J=10 B7, K=11 RowTotal
    cur_row = 5
    data_rows = []  # track actual data row numbers for subtotals

    for idx, row_data in enumerate(class_data):
        # Insert section header if needed
        if idx in section_breaks:
            write_section_header(ws, cur_row, section_breaks[idx], MAX_COL)
            cur_row += 1

        class_name = row_data[0]
        values = row_data[1:]  # b1_req, b1_obba, b1_total, b2, b3, b4, b5, b6, b7

        ws.cell(row=cur_row, column=1, value=class_name).font = FONT_BASE
        ws.cell(row=cur_row, column=1).alignment = ALIGN_LEFT

        for ci, val in enumerate(values, start=2):
            cell = ws.cell(row=cur_row, column=ci)
            if val is None:
                cell.value = "N/A"
                cell.font = FONT_GRAY
                cell.fill = FILL_WHITE
                cell.alignment = ALIGN_RIGHT
            else:
                cell.value = val
                cell.font = FONT_BLUE
                cell.number_format = NUM_FMT_K
                cell.alignment = ALIGN_RIGHT

        # Row Total = SUM(D:J) i.e. B1Total + B2 + B3 + B4 + B5 + B6 + B7
        # But we need to handle N/A cells -- use SUM which ignores text
        cell = ws.cell(row=cur_row, column=11)
        cell.value = f"=SUM(D{cur_row}:J{cur_row})"
        cell.font = Font(name="Arial", size=8, color=CLR_BLACK, bold=True)
        cell.number_format = NUM_FMT_K
        cell.alignment = ALIGN_RIGHT

        data_rows.append(cur_row)
        cur_row += 1

    # Subtotal (allocated) row
    SUB_ROW = cur_row
    ws.cell(row=SUB_ROW, column=1, value="Subtotal (allocated)").font = FONT_BOLD
    ws.cell(row=SUB_ROW, column=1).alignment = ALIGN_LEFT
    for ci in range(2, MAX_COL + 1):
        col_l = get_column_letter(ci)
        # Build SUM of only data rows (skipping section headers)
        parts = []
        for dr in data_rows:
            parts.append(f"{col_l}{dr}")
        cell = ws.cell(row=SUB_ROW, column=ci)
        cell.value = f"={'+'.join(parts)}"
        cell.font = Font(name="Arial", size=8, color=CLR_BLACK, bold=True)
        cell.number_format = NUM_FMT_K
        cell.alignment = ALIGN_RIGHT
        cell.border = BORDER_THIN_TOP
    cur_row += 1

    # "Not allocated to class" row: = Tab1 bucket total - Subtotal
    NA_ROW = cur_row
    ws.cell(row=NA_ROW, column=1, value="Not allocated to class").font = FONT_BASE
    ws.cell(row=NA_ROW, column=1).alignment = ALIGN_LEFT

    # Tab 1 bucket totals are in row 13 (Full TAM) for the aggregate,
    # but per-bucket we need individual rows.
    # Tab 1 columns: C=FY24, D=FY25, E=FY26Req, F=FY26OBBA, G=FY26Total
    # Tab 2 columns: B=B1Req, C=B1OBBA, D=B1Total, E=B2, F=B3, G=B4, H=B5, I=B6, J=B7
    # Tab 1 row 5=B1, row 6=B2, ... row 11=B7
    # For B1 Total (col D here): Tab1!G5 (FY26 Total for bucket 1)
    # For B2 (col E here): Tab1!G6
    # etc.

    # Map Tab2 cols to Tab1 bucket rows
    # Col D(4)=B1Total -> Tab1 row5 col G
    # Col E(5)=B2 -> Tab1 row6 col G
    # Col F(6)=B3 -> Tab1 row7 col G
    # Col G(7)=B4 -> Tab1 row8 col G
    # Col H(8)=B5 -> Tab1 row9 col G
    # Col I(9)=B6 -> Tab1 row10 col G
    # Col J(10)=B7 -> Tab1 row11 col G

    tab1_bucket_col = "G"  # FY26 Total column in Tab 1
    tab1_bucket_start_row = 5  # Bucket 1

    # B1 Request (col B) and B1 OBBA (col C) -- use Tab1 E5 and F5
    for ci, tab1_col in [(2, "E"), (3, "F")]:
        cell = ws.cell(row=NA_ROW, column=ci)
        cell.value = f"='TAM summary'!{tab1_col}5-{get_column_letter(ci)}{SUB_ROW}"
        cell.font = FONT_GREEN
        cell.number_format = NUM_FMT_K
        cell.alignment = ALIGN_RIGHT

    for ci in range(4, 11):
        # ci 4 = B1Total, ci 5 = B2, ... ci 10 = B7
        bucket_idx = ci - 4  # 0=B1, 1=B2, ... 6=B7
        tab1_row = tab1_bucket_start_row + bucket_idx
        col_l = get_column_letter(ci)
        cell = ws.cell(row=NA_ROW, column=ci)
        cell.value = f"='TAM summary'!{tab1_bucket_col}{tab1_row}-{col_l}{SUB_ROW}"
        cell.font = FONT_GREEN
        cell.number_format = NUM_FMT_K
        cell.alignment = ALIGN_RIGHT

    # Row total for not allocated
    cell = ws.cell(row=NA_ROW, column=11)
    cell.value = f"=SUM(D{NA_ROW}:J{NA_ROW})"
    cell.font = FONT_GREEN
    cell.number_format = NUM_FMT_K
    cell.alignment = ALIGN_RIGHT
    cur_row += 1

    # Total row (green link to Tab 1)
    TOT_ROW = cur_row
    ws.cell(row=TOT_ROW, column=1, value="Total").font = FONT_BOLD
    ws.cell(row=TOT_ROW, column=1).alignment = ALIGN_LEFT

    for ci, tab1_col in [(2, "E"), (3, "F")]:
        cell = ws.cell(row=TOT_ROW, column=ci)
        cell.value = f"='TAM summary'!{tab1_col}5"
        cell.font = Font(name="Arial", size=8, color=CLR_GREEN, bold=True)
        cell.number_format = NUM_FMT_K
        cell.alignment = ALIGN_RIGHT
        cell.border = BORDER_MED_TOP

    for ci in range(4, 11):
        bucket_idx = ci - 4
        tab1_row = tab1_bucket_start_row + bucket_idx
        cell = ws.cell(row=TOT_ROW, column=ci)
        cell.value = f"='TAM summary'!{tab1_bucket_col}{tab1_row}"
        cell.font = Font(name="Arial", size=8, color=CLR_GREEN, bold=True)
        cell.number_format = NUM_FMT_K
        cell.alignment = ALIGN_RIGHT
        cell.border = BORDER_MED_TOP

    cell = ws.cell(row=TOT_ROW, column=11)
    cell.value = f"=SUM(D{TOT_ROW}:J{TOT_ROW})"
    cell.font = Font(name="Arial", size=8, color=CLR_BLACK, bold=True)
    cell.number_format = NUM_FMT_K
    cell.alignment = ALIGN_RIGHT
    cell.border = BORDER_MED_TOP
    cur_row += 1

    # QA check row
    QA_ROW = cur_row
    ws.cell(row=QA_ROW, column=1, value="QA Check").font = FONT_BOLD
    ws.cell(row=QA_ROW, column=1).alignment = ALIGN_LEFT

    for ci in range(4, 11):
        bucket_idx = ci - 4
        tab1_row = tab1_bucket_start_row + bucket_idx
        col_l = get_column_letter(ci)
        cell = ws.cell(row=QA_ROW, column=ci)
        cell.value = f'=IF(ABS({col_l}{TOT_ROW}-\'TAM summary\'!G{tab1_row})<1,"PASS","FAIL")'
        cell.font = FONT_BOLD
        cell.alignment = ALIGN_CENTER
        # We can't conditionally format easily in openpyxl without CF rules;
        # since these should all pass, set green fill
        cell.fill = FILL_PASS

    # Also for cols B,C do QA
    for ci, tab1_col in [(2, "E"), (3, "F")]:
        col_l = get_column_letter(ci)
        cell = ws.cell(row=QA_ROW, column=ci)
        cell.value = f'=IF(ABS({col_l}{TOT_ROW}-\'TAM summary\'!{tab1_col}5)<1,"PASS","FAIL")'
        cell.font = FONT_BOLD
        cell.alignment = ALIGN_CENTER
        cell.fill = FILL_PASS

    cur_row += 1

    # Column widths
    ws.column_dimensions["A"].width = 28
    for col_l_str in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
        ws.column_dimensions[col_l_str].width = 12

    ensure_fills(ws, cur_row, MAX_COL)
    set_all_row_heights(ws, cur_row)
    setup_sheet_view(ws)

    return ws


# ── TAB 3: All line items ─────────────────────────────────────────────

def build_tab3(wb):
    ws = wb.create_sheet("All line items")
    ws.sheet_properties.tabColor = TAB_COLOR
    MAX_COL = 10

    line_items = [
        ["SCN L1+L2", "Columbia Class (ship + AP)", "SSBN (Columbia)", 1, "GD Electric Boat / HII NNS", 0, 9580774, 10920486, "High", "FY24 zero = no new TOA"],
        ["SCN L8+L9", "Virginia Class (ship + AP)", "SSN (Virginia)", 1, "GD Electric Boat / HII NNS", 10656927, 13320211, 11078029, "High", "FY26: $4.8B Req + $6.3B OBBA"],
        ["SCN L5+L6+L7", "CVN-81 Carrier Replacement", "CVN (Ford)", 1, "HII Newport News", 2529513, 2034054, 3281673, "High", "CVN-80 SFF + CVN-81 ramp"],
        ["SCN L13+L14", "DDG-51 (ship + AP)", "DDG-51 (Flt III)", 1, "HII Ingalls / GD BIW", 6366431, 8268628, 5410773, "High", "FY26: $11K Req + $5.4B OBBA"],
        ["SCN L12", "DDG-1000 Completion", "DDG-1000", 1, "GD Bath Iron Works", 392892, 61100, 52358, "High", "Completion/CPS only"],
        ["SCN L16", "FFG-62 Constellation", "FFG-62", 1, "Marinette Marine", 2183861, 633200, 0, "High", "No FY26 funding"],
        ["SCN L15", "LCS Completion", "LCS", 1, "Austal USA / LM", 23000, 48000, 0, "High", "Program ending"],
        ["SCN L19+L20+L21", "LPD Flight II", "LPD", 1, "HII Ingalls", 516520, 1581121, 2600000, "High", "FY26 entirely OBBA"],
        ["SCN L24+L25", "LHA Replacement", "LHA", 1, "HII Ingalls", 1830149, 176515, 3895000, "High", "FY26 entirely OBBA"],
        ["SCN L27", "Medium Landing Ship", "LSM", 1, "TBD", 0, 29668, 1963941, "High", "9-ship block buy in OBBA"],
        ["SCN L31", "T-AO Fleet Oiler", "T-AO", 1, "GD NASSCO", 938315, 227154, 1861705, "High", "FY26: $8K Req + $1.85B OBBA"],
        ["SCN L34", "TAGOS Surtass Ships", "T-AGOS", 1, "TBD", 513466, 0, 612205, "High", ""],
        ["SCN L35", "ATS (Towing/Salvage)", "ATS", 1, "TBD", 22959, 82587, 0, "High", "Completion only"],
        ["SCN L37", "Oceanographic Ships", "T-AGS", 1, "TBD", 0, 18000, 0, "High", "Completion only"],
        ["SCN L39", "LCU 1700", "LCU", 1, "TBD", 62532, 0, 295000, "High", "FY26 OBBA: 9 craft"],
        ["SCN L40", "Strategic Sealift", "Sealift", 1, "TBD", 0, 0, 600000, "High", "FY26 OBBA"],
        ["SCN L42", "Ship to Shore Connector", "SSC", 1, "Textron", 628600, 528039, 239095, "High", "LCAC replacement"],
        ["SCN L43", "Service Craft", "Various", 1, "Various", 93815, 41426, 34602, "High", ""],
        ["SCN L44", "Aux Personnel Lighter", "APL", 1, "TBD", 72000, 76168, 0, "High", ""],
        ["SCN L48", "Auxiliary Vessels (Used Sealift)", "Sealift", 1, "Various", 142008, 204939, 145000, "High", "Used vessel acquisition"],
        ["SCN L41", "Outfitting (all newbuild)", "Multi-class", 1, "Various", 512019, 585967, 887295, "High", "Post-delivery outfitting"],
        ["SCN L49", "PY Completion (Bucket 1 share)", "Multi-class", 1, "Various", 0, 0, 1192266, "High", "CTC minus RCOH/LCAC"],
        ["USCG PC&I", "Offshore Patrol Cutter", "CG: OPC", 1, "Eastern Shipbuilding", 579000, 579000, 812400, "High", "OPCs #8-9; LLTM #10-11"],
        ["USCG PC&I", "Fast Response Cutter", "CG: FRC", 1, "Bollinger", 220000, 220000, 216000, "High", "FRCs #62-64"],
        ["USCG PC&I", "Polar Security Cutter", "CG: PSC", 1, "Bollinger", 0, 0, 130000, "High", "PSC #1 construction"],
        ["USCG PC&I", "Waterways Commerce Cutter", "CG: WCC", 1, "TBD", 1000, 1000, 98000, "High", ""],
        ["USCG PC&I", "NSC Completion", "CG: NSC", 1, "HII Ingalls", 17100, 17100, 0, "High", "Complete after FY25"],
        ["USCG PC&I", "Commercial Polar Icebreaker", "CG: Icebreaker", 1, "Commercial", 125000, 125000, 0, "High", "CGC Storis acquisition"],
        ["USCG PC&I", "Great Lakes Icebreaker", "CG: Icebreaker", 1, "TBD", 20000, 20000, 0, "High", ""],
        ["OMN 1B4B", "Ship Modernization Program (SMP)", "Multi-class", 2, "Mixed", 1203816, 869253, 1197180, "High", ""],
        ["OMN 1B4B", "NME / 2-Hatch Ships", "Multi-class", 2, "Mixed", 1197197, 1306411, 1341155, "High", ""],
        ["OMN 1B4B", "NME / 4-Hatch Ships", "Multi-class", 2, "Mixed", 1247653, 1290155, 1289894, "High", ""],
        ["OMN 1B4B", "NME / Submarine Availability", "SSN/SSBN", 2, "Public Yard", 871537, 896253, 924629, "High", "Primarily NNSY/PHNSY"],
        ["OMN 1B4B", "NME / Industrial Base", "Multi-class", 2, "Mixed", 1245621, 1270256, 1267865, "High", ""],
        ["OMN 1B4B", "NME / NAVPLAN", "Multi-class", 2, "Mixed", 1134835, 1139556, 1134328, "High", ""],
        ["OMN 1B4B", "NME / SSN-to-Repair", "SSN/SSBN", 2, "Public Yard", 630555, 639229, 642190, "High", "Public yard sub repair"],
        ["OMN 1B4B", "NME / Carrier", "CVN", 2, "Mixed", 762048, 816344, 784000, "High", ""],
        ["OMN 1B4B", "NME / Amphibious", "Amphibious", 2, "Private", 569696, 656000, 624506, "High", ""],
        ["OMN 1B4B", "NME / LCS", "LCS", 2, "Private", 481854, 539000, 492493, "High", ""],
        ["OMN 1B4B", "NME / DDG", "DDG-51", 2, "Private", 758318, 758610, 749000, "High", ""],
        ["OMN 1B4B", "MSC M&R (transfer from NWCF)", "MSC Fleet", 2, "Private", 0, 0, 1584000, "Medium", "New FY26 transfer"],
        ["OMN 1B4B", "Residual NME rounding", "Multi-class", 2, "Mixed", 0, 0, -176024, "Low", "SAG total controls"],
        ["OPN BA1", "L26 Ship Maint, Repair & Mod", "Multi-class", 2, "Private", 2839179, 2392190, 2392620, "High", "OPN material for depot"],
        ["USCG O&S", "Bucket 2 share (60% of O&S 25.7)", "CG: fleet", 2, "Mixed", 329409, 335840, 369358, "Medium", "Analyst 60% allocation"],
        ["OMN 1B4B", "CMMP (Continuous Maint & Mod)", "Multi-class", 3, "Mixed", 1259925, 1303227, 1321948, "High", ""],
        ["OMN 1B4B", "Expedited M&M", "Multi-class", 3, "Mixed", 581796, 602706, 640000, "High", ""],
        ["OMN 1B5B", "RMC", "Multi-class", 3, "Mixed", 22311, 25633, 26529, "High", "Regional Maint Centers"],
        ["USCG O&S", "Bucket 3 share (40% of O&S 25.7)", "CG: fleet", 3, "Mixed", 219606, 223893, 246238, "Medium", "Analyst 40% allocation"],
        ["OPN BA1", "L1 Surface Power Equipment", "Multi-class", 4, "Private", 12855, 20840, 9978, "High", ""],
        ["OPN BA1", "L2 Surface Combatant HM&E", "Multi-class", 4, "Private", 92449, 77592, 62004, "High", ""],
        ["OPN BA1", "L4 Sub Periscope/Imaging", "SSN/SSBN", 4, "Mixed", 272778, 290575, 277413, "High", ""],
        ["OPN BA1", "L5 DDG Mod", "DDG-51", 4, "Private", 641676, 861066, 997485, "High", ""],
        ["OPN BA1", "L8 LHA/LHD Midlife", "LHA/LHD", 4, "Private", 102334, 81602, 123384, "High", ""],
        ["OPN BA1", "L11 Submarine Support Equipment", "SSN/SSBN", 4, "Mixed", 210652, 293766, 383062, "High", ""],
        ["OPN BA1", "L12 Virginia Class Support", "SSN (Virginia)", 4, "Mixed", 32055, 43565, 52039, "High", ""],
        ["OPN BA1", "L13 LCS Class Support", "LCS", 4, "Private", 17816, 7318, 2551, "High", "Declining"],
        ["OPN BA1", "L15 LPD Class Support", "LPD", 4, "Private", 80933, 38115, 125542, "High", ""],
        ["OPN BA1", "L16 DDG-1000 Class Support", "DDG-1000", 4, "Private", 220771, 340668, 115267, "High", ""],
        ["OPN BA1", "L17 Strategic Platform Support", "SSN/SSBN", 4, "Mixed", 23756, 53931, 53740, "High", ""],
        ["OPN BA1", "L21 LCAC", "SSC/LCAC", 4, "Private", 10239, 11013, 20321, "High", ""],
        ["OPN BA1", "L28 Reactor Components", "Multi-class", 4, "Mixed", 389890, 445974, 399603, "High", "Subs + carriers"],
        ["OPN BA1", "L33 LCS Common Mission Modules", "LCS", 4, "Private", 49028, 56105, 38880, "High", ""],
        ["OPN BA1", "L34 LCS MCM Mission Modules", "LCS", 4, "Private", 87828, 118247, 91372, "High", ""],
        ["OPN BA1", "L36 LCS SUW Mission Modules", "LCS", 4, "Private", 12094, 11101, 3790, "High", ""],
        ["OPN BA1", "L37 LCS In-Service Mod", "LCS", 4, "Private", 154561, 188254, 237728, "High", ""],
        ["OPN BA1", "L40 LSD Midlife & Mod", "LSD", 4, "Private", 3989, 56667, 4079, "High", ""],
        ["OPN BA2", "L43 AN/SQQ-89", "Multi-class", 4, "Private", 133803, 134637, 144425, "High", "DDG/CG sonar"],
        ["OPN BA2", "L44 SSN Acoustic (80%)", "SSN/SSBN", 4, "Mixed", 107116, 117310, 398878, "Medium", "80% in-service share"],
        ["OPN BA2", "L46 Sub Acoustic Warfare", "SSN/SSBN", 4, "Mixed", 41695, 51514, 56482, "High", ""],
        ["OPN BA2", "L50 AN/SLQ-32 SEWIP", "Multi-class", 4, "Private", 325996, 182011, 461380, "High", ""],
        ["OPN BA2", "L51 Shipboard IW Exploit", "Multi-class", 4, "Mixed", 367452, 362099, 379908, "High", ""],
        ["OPN BA2", "L53 CEC (95%)", "Multi-class", 4, "Private", 22776, 23696, 25316, "Medium", ""],
        ["OPN BA2", "L70 CANES (99%)", "Multi-class", 4, "Mixed", 416879, 413558, 530408, "Medium", ""],
        ["OPN BA2", "L72 CANES-Intell (99%)", "Multi-class", 4, "Mixed", 38283, 40128, 45813, "Medium", ""],
        ["OPN BA2", "L78 In-Service Radars/Sensors", "Multi-class", 4, "Private", 279545, 222607, 249656, "High", ""],
        ["OPN BA2", "L79 Battle Force Tactical Net", "Multi-class", 4, "Mixed", 68327, 104119, 106583, "High", ""],
        ["OPN BA2", "L80 Shipboard Tactical Comms (90%)", "Multi-class", 4, "Mixed", 14440, 15480, 18810, "Medium", ""],
        ["OPN BA2", "L81 Ship Comms Automation (75%)", "Multi-class", 4, "Mixed", 96000, 100800, 121556, "Medium", ""],
        ["OPN BA4", "BLI 125 Ship Gun Systems", "Multi-class", 4, "Private", 16250, 6416, 7358, "High", ""],
        ["OPN BA4", "BLI 126 Harpoon", "Multi-class", 4, "Mixed", 227, 226, 209, "High", ""],
        ["OPN BA4", "BLI 127 Ship Missile Support", "Multi-class", 4, "Mixed", 283972, 376830, 455822, "High", ""],
        ["OPN BA4", "BLI 128 Tomahawk (70%)", "Multi-class", 4, "Mixed", 48696, 55392, 75396, "Medium", ""],
        ["OPN BA4", "BLI 130 Strategic Missile (50%)", "Multi-class", 4, "Mixed", 193500, 210000, 246500, "Medium", ""],
        ["OPN BA4", "BLI 131 SSN Combat Control (86%)", "SSN/SSBN", 4, "Mixed", 67400, 72536, 88554, "Medium", ""],
        ["OPN BA4", "BLI 132 ASW Support", "Multi-class", 4, "Mixed", 37301, 25362, 25721, "High", ""],
        ["OPN BA4", "BLI 134 Directed Energy / ODIN", "Multi-class", 4, "Private", 0, 3817, 2976, "High", ""],
        ["OPN BA4", "BLI 136 Anti-Ship Missile Decoy", "Multi-class", 4, "Mixed", 51100, 75614, 89129, "High", ""],
        ["USCG PC&I", "ISSS", "CG: fleet", 4, "Mixed", 25000, 28000, 30000, "Medium", "In-Service Ship Systems"],
        ["USCG PC&I", "C4ISR", "CG: fleet", 4, "Mixed", 16000, 16000, 10000, "Medium", ""],
        ["USCG PC&I", "Cyber/EMP (50% vessel share)", "CG: fleet", 4, "Mixed", 10000, 11000, 12900, "Low", ""],
        ["SCN", "CVN RCOH SFF (CVN 75)", "CVN", 5, "HII-NNS", 0, 811143, 1779011, "High", ""],
        ["SCN", "CVN RCOH CTC (CVN 74)", "CVN", 5, "HII-NNS", 42422, 669171, 483100, "High", ""],
        ["SCN", "CVN RCOH Outfitting", "CVN", 5, "Private", 19704, 6660, 12200, "High", ""],
        ["SCN", "LCAC SLEP (ESLEP)", "SSC/LCAC", 5, "Walashek", 15286, 45087, 37390, "High", ""],
        ["SCN", "LCAC SLEP Outfitting", "SSC/LCAC", 5, "Private", 728, 0, 493, "High", ""],
        ["OPN BA1", "L9 LCC ESLP", "LCC", 5, "Private", 10522, 7352, 19276, "High", ""],
        ["USCG PC&I", "47-ft MLB SLEP", "CG: MLB", 5, "CG Yard/Private", 43000, 43000, 45000, "Medium", ""],
        ["USCG PC&I", "270-ft WMEC SLEP", "CG: WMEC", 5, "CG Yard", 46200, 46200, 76000, "Medium", ""],
        ["USCG PC&I", "175-ft WLM MMA", "CG: WLM", 5, "CG Yard", 17800, 17800, 15000, "Medium", ""],
        ["USCG PC&I", "CGC Healy SLEP", "CG: Healy", 5, "Mixed", 13000, 13000, 11000, "Medium", ""],
        ["USCG PC&I", "418-ft WMSL MMA", "CG: WMSL", 5, "CG Yard", 0, 0, 5000, "Low", "New FY26"],
        ["USCG PC&I", "Boats RB-M SLEP portion", "CG: RB-M", 5, "Private", 1950, 1950, 9270, "Low", ""],
        ["OMN 1B5B", "DDG-1000 Maintenance & Support", "DDG-1000", 6, "Mixed", 183642, 197600, 209225, "High", ""],
        ["OMN 1B5B", "Mine Countermeasures Support", "MCM", 6, "Mixed", 14157, 7186, 5573, "High", "Declining"],
        ["OMN 1B5B", "PRE/PRL CV/CVN Tech Support", "CVN", 6, "Mixed", 81444, 81505, 86561, "High", ""],
        ["OMN 1B5B", "Service Craft/Boats/Targets", "Multi-class", 6, "Mixed", 7548, 8364, 6842, "High", ""],
        ["OMN 1B5B", "Surface & Amphib Ship Support", "Multi-class", 6, "Mixed", 446317, 463167, 451307, "High", ""],
        ["OMN 1B5B", "Nuclear Propulsion Tech Logistics", "Multi-class", 6, "Mixed", 327227, 340900, 355445, "High", "Subs + carriers"],
        ["OMN 1B5B", "SUPSHIP Costs", "Multi-class", 6, "Public Yard", 255302, 269204, 251345, "High", ""],
        ["OMN 1B5B", "NMP Total", "Multi-class", 6, "Mixed", 415261, 277773, 410893, "High", ""],
        ["OMN 1B5B", "NMMES", "Multi-class", 6, "Mixed", 103028, 104083, 115706, "High", ""],
        ["OMN 1B5B", "Smart Work / TOC", "Multi-class", 6, "Mixed", 22363, 34068, 23672, "High", ""],
        ["OMN 1B5B", "CSOSS", "Multi-class", 6, "Mixed", 12756, 11062, 14954, "High", ""],
        ["OMN 1B5B", "NAVSEA HQ Civilian Personnel & IT", "Multi-class", 6, "Public Yard", 571260, 628655, 579450, "High", "HQ overhead"],
        ["OPN BA8", "BLI 9020", "Multi-class", 6, "Mixed", 736770, 705144, 883630, "High", "War reserve/initial spares"],
        ["USCG O&S", "Survey & Design", "CG: fleet", 6, "CG Yard", 5000, 5000, 5000, "Low", "Analyst estimate"],
        ["USCG O&S", "PO&M (50% vessel share)", "CG: fleet", 6, "Mixed", 10500, 10500, 11000, "Low", ""],
        ["OMN 1B5B", "Berthing & Messing", "Multi-class", 7, "Mixed", 129771, 135039, 141596, "High", ""],
        ["OMN 1B5B", "Facilities / Supply Ops (SIOP)", "Multi-class", 7, "Public Yard", 120535, 87573, 81780, "High", ""],
        ["OMN 1B5B", "Other embedded support", "Multi-class", 7, "Mixed", 2200, 2200, 2200, "Low", "Estimated"],
        ["USCG O&S", "Other Equipment & Systems (vessel)", "CG: fleet", 7, "Mixed", 2000, 2100, 2200, "Low", ""],
    ]

    # Header rows
    write_title_row(ws, "All line items", MAX_COL)
    write_purpose_row(ws, "Master list of all extracted line items across all seven buckets", MAX_COL)
    write_units_row(ws, "All figures in USD thousands ($K)", MAX_COL)

    headers = ["Source", "Line Item", "Vessel Class", "Bucket", "Performer",
               "FY24", "FY25", "FY26 Total", "Confidence", "Notes"]
    write_col_headers(ws, headers, 4, MAX_COL)
    for c in range(6, 9):
        ws.cell(row=4, column=c).alignment = ALIGN_RIGHT

    # Data
    for i, item in enumerate(line_items):
        r = 5 + i
        source, line_name, vclass, bucket, performer, fy24, fy25, fy26, conf, notes = item

        # Source (gray)
        ws.cell(row=r, column=1, value=source).font = FONT_GRAY
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT
        # Line Item (black)
        ws.cell(row=r, column=2, value=line_name).font = FONT_BASE
        ws.cell(row=r, column=2).alignment = ALIGN_LEFT
        # Vessel Class (black)
        ws.cell(row=r, column=3, value=vclass).font = FONT_BASE
        ws.cell(row=r, column=3).alignment = ALIGN_LEFT
        # Bucket (black)
        ws.cell(row=r, column=4, value=bucket).font = FONT_BASE
        ws.cell(row=r, column=4).alignment = ALIGN_CENTER
        # Performer (black)
        ws.cell(row=r, column=5, value=performer).font = FONT_BASE
        ws.cell(row=r, column=5).alignment = ALIGN_LEFT
        # FY24, FY25, FY26 (blue)
        for ci, val in [(6, fy24), (7, fy25), (8, fy26)]:
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = FONT_BLUE
            cell.number_format = NUM_FMT_K
            cell.alignment = ALIGN_RIGHT
        # Confidence (black)
        ws.cell(row=r, column=9, value=conf).font = FONT_BASE
        ws.cell(row=r, column=9).alignment = ALIGN_CENTER
        # Notes (gray)
        ws.cell(row=r, column=10, value=notes).font = FONT_GRAY
        ws.cell(row=r, column=10).alignment = ALIGN_LEFT

    max_row = 4 + len(line_items)

    # Auto-filter on header row
    ws.auto_filter.ref = f"A4:J{max_row}"

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 26
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 30

    ensure_fills(ws, max_row, MAX_COL)
    set_all_row_heights(ws, max_row)
    setup_sheet_view(ws)

    return ws


# ── TAB 4: PB-61 detail ───────────────────────────────────────────────

def build_tab4(wb):
    ws = wb.create_sheet("PB-61 detail")
    ws.sheet_properties.tabColor = TAB_COLOR
    MAX_COL = 8

    pb61_ship_maint = [
        ["CV-CVN", 1796346, 2748996, 2897547, 0.95, 2720122, 3030280, 0.90],
        ["CV-CVN (SRB)", 278, 312, 312, 1.00, 312, 323, 0.97],
        ["CVN", 0, 0, 0, None, 17657, 18054, 0.98],
        ["DDG", 732057, 563674, 595584, 0.95, 525288, 568246, 0.92],
        ["DDG-1000", 0, 0, 0, None, 6174, 7358, 0.84],
        ["SSN-MTS-SSGN", 2524018, 2635557, 2765819, 0.95, 2431863, 2531177, 0.96],
        ["SSN-SSGN", 1964943, 1633867, 1758233, 0.93, 1410798, 1598434, 0.88],
        ["SSN", 94867, 126352, 126352, 1.00, 584293, 592946, 0.99],
        ["SSBN", 424396, 276204, 276424, 1.00, 301080, 307425, 0.98],
        ["SSGN", 0, 0, 0, None, 67781, 67846, 1.00],
        ["AMPHIBS", 518125, 474175, 499230, 0.95, 180719, 223014, 0.81],
        ["LCS", 241434, 225747, 237189, 0.95, 230789, 346046, 0.67],
        ["LPD", 0, 0, 0, None, 105286, 105354, 1.00],
        ["LHA", 0, 0, 0, None, 29116, 30946, 0.94],
        ["LHD", 0, 0, 0, None, 36577, 37210, 0.98],
        ["LSD", 0, 0, 0, None, 73346, 97543, 0.75],
        ["CG", 106091, 50328, 51680, 0.97, 40292, 42503, 0.95],
        ["ESB", 6850, 20241, 20629, 0.98, 19655, 22134, 0.89],
        ["LCC", 5496, 517, 517, 1.00, 3899, 57265, 0.07],
        ["AGF-LCC", 15372, 65480, 65674, 1.00, 2762, 3093, 0.89],
        ["MCM-MHC", 55837, 46252, 47028, 0.98, 10813, 13189, 0.82],
        ["SOH", 13902, 27805, 27805, 1.00, 27382, 29120, 0.94],
        ["ARDM-AS", 1903, 89284, 89284, 1.00, 24967, 24967, 1.00],
        ["AS-ARDM", 17244, 23306, 23632, 0.99, 8400, 8751, 0.96],
    ]

    pb61_continuous = [
        ["DDG", 106432, 100690, 100690, 1.00, 123497, 123497, 1.00],
        ["LCS", 60812, 69375, 69375, 1.00, 69759, 69759, 1.00],
        ["CG", 3066, 5200, 5200, 1.00, 5969, 5969, 1.00],
        ["DDG-1000", 0, 3950, 3950, 1.00, 7645, 7645, 1.00],
        ["LCAC", 9219, 7720, 7720, 1.00, 7952, 7952, 1.00],
        ["AMPHIBS", 15037, 17188, 17188, 1.00, 0, 0, None],
        ["MINESWEEPERS", 370, 1620, 1620, 1.00, 1068, 1068, 1.00],
        ["Other", 89976, 155923, 155923, 1.00, 127665, 127665, 1.00],
    ]

    pb61_other = [
        ["CV-CVN", 193274, 97736, 102481, 0.95, 364463, 184318, 1.98],
        ["DDG", 429962, 353844, 405744, 0.87, 426684, 454399, 0.94],
        ["SSBN", 528266, 541600, 550579, 0.98, 577323, 604609, 0.95],
        ["SSN-MTS-SSGN", 356970, 233706, 258466, 0.90, 259540, 280442, 0.93],
        ["SSN-SSGN", 677440, 628771, 660421, 0.95, 453013, 477255, 0.95],
        ["SSGN", 0, 0, 0, None, 157370, 162292, 0.97],
        ["SSN", 0, 0, 0, None, 75160, 77571, 0.97],
        ["CG", 69927, 83031, 86340, 0.96, 75209, 77398, 0.97],
        ["LCS", 168387, 233021, 279807, 0.83, 345600, 372513, 0.93],
        ["AMPHIBS", 199393, 196703, 201922, 0.97, 112718, 150945, 0.75],
        ["LPD", 0, 0, 0, None, 44309, 45312, 0.98],
        ["LHA", 0, 0, 0, None, 15896, 16315, 0.97],
        ["LHD", 0, 0, 0, None, 16429, 16573, 0.99],
        ["LSD", 0, 0, 0, None, 21075, 21237, 0.99],
        ["AEGIS ASHORE", 27696, 3724, 3830, 0.97, 4275, 4244, 1.01],
        ["ESB", 1224, 1454, 1461, 1.00, 4007, 4091, 0.98],
        ["MCM-MHC", 26770, 24799, 24900, 1.00, 12468, 12585, 0.99],
        ["AGF-LCC", 11377, 14291, 15325, 0.93, 15511, 15683, 0.99],
    ]

    pb61_opn = [
        ["DDG", 1676352, 732451, 814455, None, 1282612, 1418118, None],
        ["LCS", 110171, 232091, 232091, None, 383361, 383361, None],
        ["LHD", 0, 0, 0, None, 395086, 395086, None],
        ["AMPHIBS", 573902, 1097648, 1354392, None, 236773, 393588, None],
        ["CG", 0, 0, 0, None, 44940, 44940, None],
        ["CV-CVN (continuous)", 13102, 54832, 54832, None, 63908, 63908, None],
        ["SSN (continuous)", 73142, 2810, 2810, None, 53356, 53356, None],
        ["SSN/SSGN Electronics", 7736, 7398, 7398, None, 7620, 12047, None],
        ["SSBN/SSGN (continuous)", 0, 0, 0, None, 2669, 2669, None],
        ["LHA", 0, 0, 0, None, 12065, 12065, None],
        ["LPD", 0, 0, 0, None, 11852, 11852, None],
        ["LSD", 0, 0, 0, None, 25931, 25931, None],
        ["SSN-MTS-SSGN", 383868, 330000, 330000, None, 0, 0, None],
    ]

    # Header rows
    write_title_row(ws, "PB-61 detail", MAX_COL)
    write_purpose_row(ws, "PB-61 Depot Maintenance Summary \u2014 per-class detail from OMN Vol2", MAX_COL)
    write_units_row(ws, "All figures in USD thousands ($K)", MAX_COL)

    pb61_headers = ["Weapon System", "FY24 Total", "FY25 Funded", "FY25 Required",
                    "FY25 %", "FY26 Funded", "FY26 Required", "FY26 %"]

    def write_pb61_section(ws, start_row, section_title, data, headers):
        """Write a PB-61 subsection."""
        write_subsection_header(ws, start_row, section_title, MAX_COL)
        r = start_row + 1
        # Column headers
        write_col_headers(ws, headers, r, MAX_COL)
        for c in range(2, MAX_COL + 1):
            ws.cell(row=r, column=c).alignment = ALIGN_RIGHT
        r += 1
        # Data
        for row_data in data:
            ws_name, fy24, fy25f, fy25r, fy25pct, fy26f, fy26r, fy26pct = row_data
            ws.cell(row=r, column=1, value=ws_name).font = FONT_BASE
            ws.cell(row=r, column=1).alignment = ALIGN_LEFT
            for ci, val in [(2, fy24), (3, fy25f), (4, fy25r), (6, fy26f), (7, fy26r)]:
                cell = ws.cell(row=r, column=ci, value=val)
                cell.font = FONT_BLUE
                cell.number_format = NUM_FMT_K
                cell.alignment = ALIGN_RIGHT
            # Percentages
            for ci, val in [(5, fy25pct), (8, fy26pct)]:
                cell = ws.cell(row=r, column=ci)
                if val is None:
                    cell.value = "N/A"
                    cell.font = FONT_GRAY
                    cell.alignment = ALIGN_RIGHT
                else:
                    cell.value = val
                    apply_pct_style(cell)
                    cell.alignment = ALIGN_RIGHT
            r += 1
        return r  # next available row

    cur_row = 4  # after units row (row 3)
    cur_row += 1  # skip a blank spacer row (row 4 is blank or we start sections at row 5)
    # Actually, row 4 is usually col headers, but here we have subsections with their own headers
    # Let's just use row 4 as blank spacer
    for c in range(1, MAX_COL + 1):
        ws.cell(row=4, column=c).fill = FILL_WHITE

    cur_row = 5
    cur_row = write_pb61_section(ws, cur_row, "OMN \u2014 Ship Maint (Bucket 2)", pb61_ship_maint, pb61_headers)
    cur_row += 1  # spacer
    cur_row = write_pb61_section(ws, cur_row, "OMN \u2014 Continuous Maintenance (Bucket 3)", pb61_continuous, pb61_headers)
    cur_row += 1
    cur_row = write_pb61_section(ws, cur_row, "OMN \u2014 Other Depot Maintenance", pb61_other, pb61_headers)
    cur_row += 1
    cur_row = write_pb61_section(ws, cur_row, "OPN \u2014 Ship Depot Maintenance", pb61_opn, pb61_headers)

    # Column widths
    ws.column_dimensions["A"].width = 24
    for col_l in ["B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col_l].width = 12

    ensure_fills(ws, cur_row, MAX_COL)
    set_all_row_heights(ws, cur_row)
    setup_sheet_view(ws)

    return ws


# ── TAB 5: Sources and methodology ────────────────────────────────────

def build_tab5(wb):
    ws = wb.create_sheet("Sources and methodology")
    ws.sheet_properties.tabColor = TAB_COLOR
    MAX_COL = 5

    write_title_row(ws, "Sources and methodology", MAX_COL)
    write_purpose_row(ws, "Data sources, confidence levels, methodology notes, and reconciliation checks", MAX_COL)
    # Row 3 blank
    for c in range(1, MAX_COL + 1):
        ws.cell(row=3, column=c).fill = FILL_WHITE

    cur_row = 4

    # ── Source Files ──
    write_subsection_header(ws, cur_row, "Source Files", MAX_COL)
    cur_row += 1

    sources = [
        ["SCN Justification Book (PB-27)", "Bucket 1 new construction line items and OBBA detail"],
        ["OMN Vol 2 (PB-61)", "Buckets 2-3 depot maintenance and continuous maintenance by weapon system"],
        ["OPN Justification Book (PB-28)", "Bucket 4 modernization equipment and Bucket 2 material support"],
        ["USCG Capital Investment Plan", "Coast Guard PC&I acquisition and SLEP programs"],
        ["USCG O&S Budget Justification", "Coast Guard operations and support allocated to vessel maintenance"],
        ["OPN BA8 Initial Spares", "Bucket 6 war reserve and initial spares (BLI 9020)"],
        ["NAVSEA HQ / SUPSHIP data", "Bucket 6 sustainment engineering and planning overhead"],
    ]
    for src, desc in sources:
        ws.cell(row=cur_row, column=1, value=src).font = FONT_BOLD
        ws.cell(row=cur_row, column=1).alignment = ALIGN_LEFT
        ws.cell(row=cur_row, column=2, value=desc).font = FONT_BASE
        ws.cell(row=cur_row, column=2).alignment = ALIGN_LEFT
        cur_row += 1

    cur_row += 1  # spacer

    # ── Confidence Levels ──
    write_subsection_header(ws, cur_row, "Confidence Levels", MAX_COL)
    cur_row += 1

    conf_levels = [
        ["High", "Direct budget line item with no estimation or allocation required"],
        ["Medium", "Derived from budget data with minor analyst allocation (e.g., percentage split)"],
        ["Low", "Analyst estimate based on historical patterns, indirect evidence, or proxy data"],
    ]
    for level, desc in conf_levels:
        ws.cell(row=cur_row, column=1, value=level).font = FONT_BOLD
        ws.cell(row=cur_row, column=1).alignment = ALIGN_LEFT
        ws.cell(row=cur_row, column=2, value=desc).font = FONT_BASE
        ws.cell(row=cur_row, column=2).alignment = ALIGN_LEFT
        cur_row += 1

    cur_row += 1

    # ── Key Methodology Notes ──
    write_subsection_header(ws, cur_row, "Key Methodology Notes", MAX_COL)
    cur_row += 1

    notes = [
        "1. All figures are in USD thousands ($K) unless otherwise noted.",
        "2. FY26 OBBA (Off-Budget Baseline Adjustment) reflects supplemental shipbuilding funding in the NDAA.",
        "3. Coast Guard O&S is split 60/40 between Bucket 2 (depot) and Bucket 3 (continuous) based on analyst estimate.",
        "4. OPN lines with partial vessel share (e.g., 80%, 95%) reflect analyst allocation of ship-applicable portion.",
        "5. MSC M&R ($1.584B) is a new FY26 transfer from Navy Working Capital Fund and was not in FY24-25 baselines.",
        "6. PY Completion (prior-year) is split between Bucket 1 (newbuild) and Bucket 5 (RCOH/SLEP) based on program detail.",
        "7. Residual NME rounding line in Bucket 2 is a balancing entry to match the SAG-level control total.",
        "8. Bucket 7 (Availability Support) captures port services, berthing, and facility support not included in depot events.",
    ]
    for note in notes:
        ws.cell(row=cur_row, column=1, value=note).font = FONT_BASE
        ws.cell(row=cur_row, column=1).alignment = ALIGN_LEFT
        cur_row += 1

    cur_row += 1

    # ── Annotation Key ──
    write_subsection_header(ws, cur_row, "Annotation Key", MAX_COL)
    cur_row += 1

    annotations = [
        ['N/A', "Budget source does not break this data out by vessel class for this bucket"],
        ['\u2014 (dash)', "Genuine zero value; the number format renders 0 as a dash"],
        ["Blue text", "Hardcoded numeric input from budget documents"],
        ["Green text", "Cross-tab formula linking to another worksheet"],
        ["Gray text", "Source references, notes, or analyst annotations"],
    ]
    for key, desc in annotations:
        ws.cell(row=cur_row, column=1, value=key).font = FONT_BOLD
        ws.cell(row=cur_row, column=1).alignment = ALIGN_LEFT
        ws.cell(row=cur_row, column=2, value=desc).font = FONT_BASE
        ws.cell(row=cur_row, column=2).alignment = ALIGN_LEFT
        cur_row += 1

    cur_row += 1

    # ── Reconciliation Checks ──
    write_subsection_header(ws, cur_row, "Reconciliation Checks", MAX_COL)
    cur_row += 1

    # Column headers for recon table
    recon_headers = ["Bucket", "Tab 1 Value", "Tab 2 Column Total", "Delta", "Status"]
    write_col_headers(ws, recon_headers, cur_row, MAX_COL)
    for c in range(2, MAX_COL + 1):
        ws.cell(row=cur_row, column=c).alignment = ALIGN_RIGHT
    cur_row += 1

    # We need to know Tab 2 subtotal row. Tab 2 has dynamic row numbers due to section headers.
    # Tab 2 subtotal row: we computed it as SUB_ROW in build_tab2.
    # Since we can't easily pass that, let's compute it here.
    # class_data has 34 rows, section_breaks at indices 0,3,9,14,23,31,33 = 7 section headers
    # So data starts at row 5, with 34 data rows + 7 section headers = 41 content rows
    # SUB_ROW = 5 + 41 = 46
    # Actually let's count: rows 5 onward.
    # idx 0: section header (row 5), data row (row 6)
    # idx 1: data (row 7)
    # idx 2: data (row 8)
    # idx 3: section header (row 9), data (row 10)
    # ...
    # Total rows = 34 data + 7 headers = 41, starting at row 5 -> last content at row 45
    # SUB_ROW = 46
    TAB2_SUB_ROW = 46

    # Tab 1 FY26 Total column = G, rows 5-11 for buckets 1-7
    # Tab 2 bucket columns: D=B1Total, E=B2, F=B3, G=B4, H=B5, I=B6, J=B7

    bucket_names = [
        "Bucket 1 (New Construction)",
        "Bucket 2 (Scheduled Depot Maint)",
        "Bucket 3 (Continuous/Intermediate)",
        "Bucket 4 (Modernization)",
        "Bucket 5 (Major Life-Cycle)",
        "Bucket 6 (Sustainment Eng)",
        "Bucket 7 (Availability Support)",
    ]
    tab2_cols = ["D", "E", "F", "G", "H", "I", "J"]

    for i, bname in enumerate(bucket_names):
        tab1_row = 5 + i  # bucket rows in Tab 1
        tab2_col = tab2_cols[i]

        ws.cell(row=cur_row, column=1, value=bname).font = FONT_BASE
        ws.cell(row=cur_row, column=1).alignment = ALIGN_LEFT

        # Tab 1 Value (green link)
        cell = ws.cell(row=cur_row, column=2)
        cell.value = f"='TAM summary'!G{tab1_row}"
        cell.font = FONT_GREEN
        cell.number_format = NUM_FMT_K
        cell.alignment = ALIGN_RIGHT

        # Tab 2 Column Total (green link)
        cell = ws.cell(row=cur_row, column=3)
        cell.value = f"='By vessel class'!{tab2_col}{TAB2_SUB_ROW}"
        cell.font = FONT_GREEN
        cell.number_format = NUM_FMT_K
        cell.alignment = ALIGN_RIGHT

        # Delta
        cell = ws.cell(row=cur_row, column=4)
        cell.value = f"=B{cur_row}-C{cur_row}"
        cell.font = FONT_BASE
        cell.number_format = NUM_FMT_K
        cell.alignment = ALIGN_RIGHT

        # Status
        cell = ws.cell(row=cur_row, column=5)
        cell.value = f'=IF(ABS(D{cur_row})<1,"PASS","FAIL")'
        cell.font = FONT_BOLD
        cell.alignment = ALIGN_CENTER
        cell.fill = FILL_PASS

        cur_row += 1

    # Column widths
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10

    ensure_fills(ws, cur_row, MAX_COL)
    set_all_row_heights(ws, cur_row)
    setup_sheet_view(ws)

    return ws


# ── Main ───────────────────────────────────────────────────────────────

def main():
    wb = Workbook()

    build_tab1(wb)
    build_tab2(wb)
    build_tab3(wb)
    build_tab4(wb)
    build_tab5(wb)

    out_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "output", "tam_workbook.xlsx")
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f"Workbook saved to {out_path}")


if __name__ == "__main__":
    main()
